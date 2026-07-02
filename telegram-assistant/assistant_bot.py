"""
assistant_bot.py — Персональный AI-ассистент Александра
"""
import asyncio
import json
import logging
import sys
import tempfile
import os
from datetime import datetime
from zoneinfo import ZoneInfo
from pathlib import Path

import re

_MSK = ZoneInfo("Europe/Moscow")

sys.path.insert(0, "/home/parser/bots/assistant")


_MAX_TOKENS = 2000
_MAX_HISTORY_TOKENS = 20_000  # ~80 КБ текста; защита от дорогих fallback-моделей


def _count_history_tokens(history: list) -> int:
    """Грубая оценка токенов истории: 1 токен ≈ 4 символа."""
    total = 0
    for msg in history:
        c = msg.get("content", "")
        if isinstance(c, str):
            total += len(c)
        elif isinstance(c, list):
            for part in c:
                total += len(str(part))
    return total // 4

# Один или больше пайпов с опциональными пробелами — ASCII | (U+007C) и полноширинный ｜ (U+FF5C)
# Важно: (?:...) non-capturing, иначе group-индексы в parse-функции сдвигаются
_D = r'(?:[|｜]\s*)+'

_DSML_CLOSED_RE = re.compile(rf'<\s*{_D}DSML\s*{_D}tool_calls\s*>.*?</\s*{_D}DSML\s*{_D}tool_calls\s*>', re.DOTALL)
_DSML_OPEN_RE   = re.compile(rf'<\s*{_D}DSML\s*{_D}tool_calls\s*>.*', re.DOTALL)
_DSML_TAG_RE    = re.compile(rf'<\s*/?\s*{_D}DSML\s*{_D}[^>]*>', re.DOTALL)
_DSML_CHECK_RE  = re.compile(r'DSML', re.IGNORECASE)


def _strip_dsml(text: str) -> str:
    text = _DSML_CLOSED_RE.sub('', text)   # убрать закрытые блоки
    text = _DSML_OPEN_RE.sub('', text)      # убрать незакрытые блоки
    text = _DSML_TAG_RE.sub('', text)       # убрать одиночные теги-остатки
    return text.strip()


def _has_dsml(text: str) -> bool:
    return bool(_DSML_CHECK_RE.search(text))


def _clean_response(text: str) -> str:
    text = _strip_dsml(text)
    text = re.sub(r'<function_calls>.*?</function_calls>', '', text, flags=re.DOTALL)
    text = re.sub(r'<invoke\b.*?</invoke>', '', text, flags=re.DOTALL)
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text.strip() or "(пустой ответ)"


# ── Проектные заметки → Google Sheets ─────────────────────────────
_PROJECT_NOTES_SHEET_ID = "1aiCKUs-Le-adHSfAOOC2AHRs1vCWzm-E3lucPTbRRkc"
_PROJECT_SA = "/home/parser/config/personal/service_account.json"
_PROJECT_MAP = {
    "губаха": "Губаха", "губахи": "Губаха", "губахе": "Губаха", "губаху": "Губаха",
    "юникорн": "Юникорн", "unicorn": "Юникорн",
    "монблан": "Монблан", "монблана": "Монблан", "монблане": "Монблан",
    "прочее": "Без проекта", "проче": "Без проекта", "офис": "Без проекта",
    "разное": "Без проекта", "общее": "Без проекта", "без проекта": "Без проекта",
}


async def _save_project_note(project_raw: str, text: str) -> str:
    """Асинхронная обёртка — запускает синхронный Sheets API в executor."""
    loop = asyncio.get_running_loop()
    return await loop.run_in_executor(None, _save_project_note_sync, project_raw, text)


def _save_project_note_sync(project_raw: str, text: str) -> str:
    from google.oauth2.service_account import Credentials
    from googleapiclient.discovery import build
    from datetime import datetime as _dt

    key = project_raw.strip().lower()
    project = _PROJECT_MAP.get(key, project_raw.strip().capitalize())

    now = _dt.now()
    row = [now.strftime("%d.%m.%Y"), now.strftime("%H:%M"), text.strip()]

    try:
        creds = Credentials.from_service_account_file(
            _PROJECT_SA, scopes=["https://www.googleapis.com/auth/spreadsheets"]
        )
        svc = build("sheets", "v4", credentials=creds, cache_discovery=False)

        # Проверяем — есть ли лист, если нет — создаём
        meta = svc.spreadsheets().get(spreadsheetId=_PROJECT_NOTES_SHEET_ID).execute()
        existing = [s["properties"]["title"] for s in meta["sheets"]]
        if project not in existing:
            svc.spreadsheets().batchUpdate(
                spreadsheetId=_PROJECT_NOTES_SHEET_ID,
                body={"requests": [{"addSheet": {"properties": {"title": project}}}]}
            ).execute()
            svc.spreadsheets().values().update(
                spreadsheetId=_PROJECT_NOTES_SHEET_ID,
                range=f"'{project}'!A1:C1",
                valueInputOption="USER_ENTERED",
                body={"values": [["Дата", "Время", "Мысль / Идея / План"]]}
            ).execute()

        svc.spreadsheets().values().append(
            spreadsheetId=_PROJECT_NOTES_SHEET_ID,
            range=f"'{project}'!A:C",
            valueInputOption="USER_ENTERED",
            insertDataOption="INSERT_ROWS",
            body={"values": [row]},
        ).execute()
        return f"✅ Записано в *{project}*:\n_{text.strip()}_"
    except Exception as e:
        return f"❌ Ошибка записи в Sheets: {e}"


def _parse_dsml_tool_calls(text: str) -> list[tuple[str, dict]]:
    """Парсит DSML-формат тул-коллов из текста ответа модели."""
    if not _DSML_CHECK_RE.search(text):
        return []
    invoke_re = re.compile(
        rf'<\s*{_D}DSML\s*{_D}invoke\s+name="([^"]+)"[^>]*>(.*?)</\s*{_D}DSML\s*{_D}invoke\s*>',
        re.DOTALL,
    )
    param_re = re.compile(
        rf'<\s*{_D}DSML\s*{_D}parameter\s+name="([^"]+)"[^>]*>(.*?)</\s*{_D}DSML\s*{_D}parameter\s*>',
        re.DOTALL,
    )
    calls = []
    for inv_m in invoke_re.finditer(text):
        name = inv_m.group(1)
        args = {m.group(1): m.group(2).strip() for m in param_re.finditer(inv_m.group(2))}
        calls.append((name, args))
    return calls

from aiogram import Bot, Dispatcher, F
from aiogram.filters import Command
from aiogram.types import Message, CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton
from apscheduler.schedulers.asyncio import AsyncIOScheduler
import openai

import config
from tools.db import init_db
from tools import notes, reminders as rem_tool, calendar, web, memory as mem_tool, team_tasks, email_tool, contacts as contacts_tool, vault as vault_tool

# ── Инициализация ─────────────────────────────────────────────
bot = Bot(token=config.TELEGRAM_TOKEN)
dp = Dispatcher()
scheduler = AsyncIOScheduler(timezone="Europe/Moscow")

ai_client = openai.AsyncOpenAI(
    base_url=config.ROUTERAI_BASE_URL,
    api_key=config.ROUTERAI_API_KEY,
    timeout=90.0,  # 90 сек максимум; без таймаута бот зависает навсегда
)

# RouterAI иногда подменяет запрошенную модель на дорогую (например opus) и списывает деньги.
# Оборачиваем create() для всех вызовов: при подмене на неразрешённую модель — переключаемся
# на следующую из списка, не падаем с ошибкой и не показываем пользователю сбой.
_FALLBACK_MODELS = [
    'deepseek/deepseek-v4-pro',
    'anthropic/claude-sonnet-4.6',
    'perplexity/sonar-pro',
    'qwen/qwen3-next-80b-a3b-thinking',
]
_original_create = ai_client.chat.completions.create


async def _create_with_fallback(*, model: str, **kwargs):
    models_to_try = [model] + [m for m in _FALLBACK_MODELS if m != model]
    last_exc = None
    for m in models_to_try:
        try:
            resp = await _original_create(model=m, **kwargs)
        except Exception as e:
            last_exc = e
            continue
        actual_model = getattr(resp, 'model', '') or ''
        if actual_model and not any(allowed in actual_model for allowed in _FALLBACK_MODELS):
            logging.warning("RouterAI подменил %s на %s — переключаюсь на следующую модель", m, actual_model)
            continue
        return resp
    raise RuntimeError(f"Все разрешённые модели недоступны: {last_exc}")


ai_client.chat.completions.create = _create_with_fallback

histories: dict[int, list[dict]] = {}
_pending_call_summaries: dict[int, str] = {}  # саммари звонков ожидающие действия
_pending_call_transcripts: dict[int, str] = {}  # транскрипции для повторной генерации
_awaiting_call_correction: dict[int, int] = {}  # user_id → message_id саммари
_pending_reformat: dict[int, dict] = {}  # {user_id: {instruction, original}}
_last_message_time: dict[int, float] = {}
_CONTEXT_TTL = 1800  # 30 минут — после паузы контекст считается новым

# ── База знаний ───────────────────────────────────────────────
def _load_knowledge() -> str:
    knowledge_dir = Path(getattr(config, "KNOWLEDGE_DIR", "/home/parser/bots/assistant/knowledge"))
    parts = []
    for fname in ["projects.md", "user.md", "bot_facts.md"]:
        fpath = knowledge_dir / fname
        if fpath.exists():
            parts.append(fpath.read_text(encoding="utf-8"))
    return "\n\n---\n\n".join(parts) if parts else ""

KNOWLEDGE = _load_knowledge()

# ── Инструменты LLM ───────────────────────────────────────────
TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "get_team_info",
            "description": (
                "Получить точную информацию о команде ВК Губаха: имена, должности, зоны ответственности. "
                "Вызывай ПЕРВЫМ при ЛЮБОМ вопросе о Викторе, Евгении, Надежде, Управляющем, Тех.директоре. "
                "НЕ ищи этих людей через web_search — они сотрудники Александра, их данные только здесь."
            ),
            "parameters": {"type": "object", "properties": {}},
        },
    },
    {
        "type": "function",
        "function": {
            "name": "web_search",
            "description": (
                "Поиск во внешнем интернете. "
                "Используй ТОЛЬКО для: погода, курсы валют, внешние новости, публичные факты из открытых источников. "
                "ЗАПРЕЩЕНО использовать для: вопросов о команде Александра (Виктор, Евгения, Надежда, Управляющий, Тех.директор), "
                "его проектах (Губаха, Монблан, Юникорн), KPI, задачах, расписании, контактах, напоминаниях. "
                "Для всего перечисленного используй данные из системного промпта или recall_facts."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {"type": "string", "description": "Поисковый запрос"}
                },
                "required": ["query"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "fetch_url",
            "description": "Загрузить и прочитать содержимое веб-страницы по URL.",
            "parameters": {
                "type": "object",
                "properties": {
                    "url": {"type": "string", "description": "Полный URL страницы"}
                },
                "required": ["url"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "add_note",
            "description": "Сохранить заметку.",
            "parameters": {
                "type": "object",
                "properties": {
                    "text": {"type": "string"},
                    "tags": {"type": "string", "description": "Теги через запятую (необязательно)"},
                },
                "required": ["text"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "list_notes",
            "description": "Показать последние заметки.",
            "parameters": {
                "type": "object",
                "properties": {
                    "limit": {"type": "integer"}
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "search_notes",
            "description": "Найти заметки по ключевому слову.",
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {"type": "string"}
                },
                "required": ["query"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "delete_note",
            "description": "Удалить заметку по ID.",
            "parameters": {
                "type": "object",
                "properties": {
                    "note_id": {"type": "integer"}
                },
                "required": ["note_id"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "add_reminder",
            "description": "Установить напоминание. Время в ISO8601: 2026-05-20T18:00:00. Для повторяющихся — передай recurrence.",
            "parameters": {
                "type": "object",
                "properties": {
                    "text": {"type": "string"},
                    "remind_at": {"type": "string", "description": "ISO8601, например 2026-05-21T09:00:00 — первое срабатывание"},
                    "recurrence": {
                        "type": "string",
                        "enum": ["daily", "weekly", "monthly", "yearly"],
                        "description": "Периодичность: daily=каждый день, weekly=каждую неделю, monthly=каждый месяц, yearly=каждый год. Не передавай если разовое напоминание.",
                    },
                },
                "required": ["text", "remind_at"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "list_reminders",
            "description": "Показать активные напоминания.",
            "parameters": {"type": "object", "properties": {}},
        },
    },
    {
        "type": "function",
        "function": {
            "name": "cancel_reminder",
            "description": "Отменить напоминание по ID.",
            "parameters": {
                "type": "object",
                "properties": {
                    "reminder_id": {"type": "integer"}
                },
                "required": ["reminder_id"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "get_schedule",
            "description": (
                "Показать расписание на конкретный день: события из Google Calendar + напоминания. "
                "Используй при ЛЮБОМ вопросе о планах, делах или расписании на день — "
                "сегодня, завтра, в конкретную дату или день недели. "
                "Сам переведи дату из слов в ISO-формат YYYY-MM-DD."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "date": {"type": "string", "description": "Дата в формате ISO: 2026-05-26"}
                },
                "required": ["date"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "get_calendar_events",
            "description": "Получить события из Google Calendar на несколько дней вперёд (без напоминаний). Используй только если нужен диапазон дней, а не конкретная дата.",
            "parameters": {
                "type": "object",
                "properties": {
                    "days": {"type": "integer", "description": "На сколько дней вперёд (по умолчанию 7)"}
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "create_calendar_event",
            "description": "Создать событие в Google Calendar.",
            "parameters": {
                "type": "object",
                "properties": {
                    "title": {"type": "string"},
                    "start": {"type": "string", "description": "ISO8601, например 2026-05-21T10:00:00"},
                    "end": {"type": "string", "description": "ISO8601 (необязательно, по умолчанию +1 час)"},
                    "description": {"type": "string"},
                },
                "required": ["title", "start"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "remember_fact",
            "description": "Запомнить важный факт о пользователе, проекте или договорённости. Используй когда пользователь сообщает что-то важное о себе, своих предпочтениях или проектах.",
            "parameters": {
                "type": "object",
                "properties": {
                    "text": {"type": "string", "description": "Факт для запоминания"},
                    "category": {
                        "type": "string",
                        "enum": ["fact", "preference", "project", "decision"],
                        "description": "fact=общий факт, preference=предпочтение, project=о проекте, decision=принятое решение",
                    },
                },
                "required": ["text"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "recall_facts",
            "description": (
                "Найти факты из личной памяти по теме. "
                "Вызывай ПЕРВЫМ при любом вопросе о людях, проектах или договорённостях Александра — "
                "до того как использовать web_search. "
                "Например: кто такой Виктор, что помнишь про Евгению, детали по проекту."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {"type": "string", "description": "Тема или ключевые слова для поиска"}
                },
                "required": ["query"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "list_memories",
            "description": "Показать всё что бот помнит, по категории.",
            "parameters": {
                "type": "object",
                "properties": {
                    "category": {
                        "type": "string",
                        "enum": ["fact", "preference", "project", "decision", ""],
                        "description": "Категория или пустая строка для всех",
                    }
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "forget_fact",
            "description": "Удалить факт из памяти по ID.",
            "parameters": {
                "type": "object",
                "properties": {
                    "memory_id": {"type": "integer"}
                },
                "required": ["memory_id"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "update_knowledge",
            "description": (
                "Обновить базу знаний о проектах или пользователе. "
                "Вызывай когда пользователь сообщает важное изменение в проекте, "
                "новый проект, смену роли, решение по архитектуре — то что должно "
                "сохраниться навсегда и быть доступно в будущих сессиях."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "target": {
                        "type": "string",
                        "enum": ["projects", "user", "bot_facts"],
                        "description": "'projects' — изменения в проектах, 'user' — информация об Александре, 'bot_facts' — запомненные факты из диалогов",
                    },
                    "content": {
                        "type": "string",
                        "description": "Текст для добавления в markdown-файл (можно использовать заголовки ##)",
                    },
                },
                "required": ["target", "content"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "restart_bot",
            "description": "Перезапускает бота на сервере через systemctl. Использовать когда нужно применить изменения в конфигурации или коде.",
            "parameters": {"type": "object", "properties": {}, "required": []},
        },
    },
    {
        "type": "function",
        "function": {
            "name": "add_contact",
            "description": (
                "Сохранить или обновить контакт (имя, телефон, Telegram, email). "
                "Используй когда пользователь говорит 'запомни контакт', 'добавь контакт' или называет данные человека."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "name":     {"type": "string", "description": "Имя и фамилия"},
                    "phone":    {"type": "string", "description": "Номер телефона"},
                    "telegram": {"type": "string", "description": "Telegram username (с @ или без)"},
                    "email":    {"type": "string", "description": "Email адрес"},
                    "notes":    {"type": "string", "description": "Заметки о человеке"},
                },
                "required": ["name"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "find_contact",
            "description": "Найти контакт по имени, телефону или Telegram. Возвращает карточку контакта.",
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {"type": "string", "description": "Имя или часть имени, телефон, @username"}
                },
                "required": ["query"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "list_contacts",
            "description": "Показать все сохранённые контакты.",
            "parameters": {"type": "object", "properties": {}},
        },
    },
    {
        "type": "function",
        "function": {
            "name": "delete_contact",
            "description": "Удалить контакт по ID.",
            "parameters": {
                "type": "object",
                "properties": {
                    "contact_id": {"type": "integer"}
                },
                "required": ["contact_id"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "get_contact_card",
            "description": (
                "Получить карточку контакта для пересылки другому человеку. "
                "Используй когда просят 'перешли контакт', 'поделись контактом'."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {"type": "string", "description": "Имя контакта"}
                },
                "required": ["query"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "search_files",
            "description": (
                "Найти файл на Mac Александра по имени или ключевым словам. "
                "Используй когда просят найти файл, документ, PDF, отчёт и т.п. "
                "Возвращает список найденных файлов с путями."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {"type": "string", "description": "Имя файла или ключевые слова"}
                },
                "required": ["query"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "send_file_email",
            "description": (
                "Отправить файл с Mac на email. Вызывай после того как пользователь подтвердил "
                "какой файл и кому отправить. Принимает полный путь к файлу из search_files. "
                "Известные контакты: Катя, Виктор, Алексей (рабочая), Алексей (личная)."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "file_path": {"type": "string", "description": "Полный путь к файлу (из результата search_files)"},
                    "to_name": {"type": "string", "description": "Имя получателя из адресной книги или email напрямую"},
                    "subject": {"type": "string", "description": "Тема письма (необязательно, по умолчанию — имя файла)"},
                },
                "required": ["file_path", "to_name"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "check_send_status",
            "description": "Проверить статус последней отправки файла по email.",
            "parameters": {"type": "object", "properties": {}},
        },
    },
    {
        "type": "function",
        "function": {
            "name": "add_team_task",
            "description": (
                "Добавить задачу для члена команды ВК Губаха в таблицу 'Задачи недели'. "
                "Команда: Виктор, Евгения, Надежда, Управляющий, Тех.директор. "
                "Используй когда говорят 'поставь задачу', 'скажи Евгении', 'добавь в задачи'."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "executor": {
                        "type": "string",
                        "description": "Исполнитель: Виктор, Евгения, Надежда, Управляющий или Тех.директор",
                    },
                    "task": {
                        "type": "string",
                        "description": "Описание задачи",
                    },
                    "deadline": {
                        "type": "string",
                        "description": "Срок в любом формате, например 'пятница', '30.05', '2026-05-30'",
                    },
                    "result": {
                        "type": "string",
                        "description": "Ожидаемый результат (необязательно)",
                    },
                    "how_to_check": {
                        "type": "string",
                        "description": "Как проверить выполнение (необязательно)",
                    },
                    "block": {
                        "type": "string",
                        "description": "Блок A или B (по умолчанию A)",
                    },
                },
                "required": ["executor", "task"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "save_card",
            "description": (
                "Сохранить идею, заметку, решение, вывод в Google Sheets или vault. "
                "Для card_type='idea': пишет в Google Sheets (таблица проектных заметок). "
                "Используй когда: 'заметка для проекта', 'идея для Губахи', 'запомни идею', 'сохрани мысль'."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "card_type": {
                        "type": "string",
                        "description": "Тип: idea (→ Google Sheets), learning, decision, note, contact_card",
                    },
                    "title": {
                        "type": "string",
                        "description": "Краткий заголовок (3-7 слов)",
                    },
                    "content": {
                        "type": "string",
                        "description": "Содержание",
                    },
                    "project": {
                        "type": "string",
                        "description": "Проект для idea: Губаха, Монблан, Юникорн, Без проекта (по умолчанию)",
                    },
                    "tags": {
                        "type": "string",
                        "description": "Теги через запятую, например: 'ai, автоматизация, проект'",
                    },
                },
                "required": ["card_type", "title", "content"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "search_vault",
            "description": (
                "Поиск по базе знаний (vault): идеи, решения, выводы, заметки. "
                "Используй когда: 'найди идею про X', 'что я думал о Y', 'поищи в базе знаний'."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "Поисковый запрос",
                    },
                },
                "required": ["query"],
            },
        },
    },
]

# ── Выполнение инструментов ───────────────────────────────────
def execute_tool(name: str, args: dict, user_id: int) -> str:
    try:
        if name == "get_team_info":
            from pathlib import Path as _Path
            import re as _re
            knowledge_dir = _Path(getattr(config, "KNOWLEDGE_DIR", "/home/parser/bots/assistant/knowledge"))
            fpath = knowledge_dir / "projects.md"
            if fpath.exists():
                content = fpath.read_text(encoding="utf-8")
                match = _re.search(r'(Команда ВК Губаха.*?)(?=\n- Стратегические|\n###|\Z)', content, _re.DOTALL)
                if match:
                    return match.group(1).strip()
            return "Информация о команде не найдена в базе знаний."
        elif name == "web_search":
            return web.search(args["query"])
        elif name == "fetch_url":
            return web.fetch_url(args["url"])
        elif name == "add_note":
            return notes.add_note(args["text"], args.get("tags", ""))
        elif name == "list_notes":
            return notes.list_notes(args.get("limit", 10))
        elif name == "search_notes":
            return notes.search_notes(args["query"])
        elif name == "delete_note":
            return notes.delete_note(args["note_id"])
        elif name == "add_reminder":
            remind_at = args.get("remind_at")
            if not remind_at:
                from datetime import timedelta as _td
                now_msk = datetime.now(_MSK)
                tomorrow = (now_msk + _td(days=1)).strftime("%Y-%m-%d")
                today = now_msk.strftime("%Y-%m-%d")
                return (
                    f"ОШИБКА: поле remind_at не передано. "
                    f"Текущее время МСК: {now_msk.strftime('%Y-%m-%dT%H:%M:%S')}. "
                    f"Сегодня={today}, завтра={tomorrow}. "
                    f"Вызови add_reminder снова с remind_at в формате YYYY-MM-DDTHH:MM:SS, "
                    f"например remind_at='{tomorrow}T09:30:00' для завтра в 9:30."
                )
            return rem_tool.add_reminder(args["text"], remind_at, user_id, args.get("recurrence"))
        elif name == "list_reminders":
            return rem_tool.list_reminders(user_id)
        elif name == "cancel_reminder":
            return rem_tool.cancel_reminder(args["reminder_id"])
        elif name == "get_schedule":
            date_iso = args["date"]
            events = calendar.get_schedule_for_date(config.GOOGLE_CALENDAR_ID, config.SERVICE_ACCOUNT_JSON, date_iso)
            reminders = rem_tool.get_reminders_for_date(user_id, date_iso)
            lines = []
            if events:
                lines.append("📅 Календарь:")
                for ev in events:
                    start = ev["start"].get("dateTime", ev["start"].get("date", ""))
                    prefix = datetime.fromisoformat(start).strftime("%H:%M") if "T" in start else "весь день"
                    lines.append(f"• {prefix} — {ev.get('summary', '(без названия)')}")
            else:
                lines.append("📅 Календарь: событий нет")
            if reminders:
                lines.append("⏰ Напоминания:")
                for r in reminders:
                    lines.append(f"• {r['remind_at'][11:16]} — {r['text']}")
            return "\n".join(lines) if lines else "На этот день ничего не запланировано."
        elif name == "get_calendar_events":
            return calendar.get_events(config.GOOGLE_CALENDAR_ID, config.SERVICE_ACCOUNT_JSON, args.get("days", 7))
        elif name == "create_calendar_event":
            return calendar.create_event(
                config.GOOGLE_CALENDAR_ID,
                config.SERVICE_ACCOUNT_JSON,
                args["title"],
                args["start"],
                args.get("end"),
                args.get("description", ""),
            )
        elif name == "remember_fact":
            return mem_tool.remember_fact(args["text"], args.get("category", "fact"))
        elif name == "recall_facts":
            return mem_tool.recall_facts(args["query"])
        elif name == "list_memories":
            return mem_tool.list_memories(args.get("category", ""))
        elif name == "forget_fact":
            return mem_tool.forget_fact(args["memory_id"])
        elif name == "update_knowledge":
            return mem_tool.update_knowledge(args["target"], args["content"])
        elif name == "restart_bot":
            import subprocess
            subprocess.Popen(
                ["systemctl", "restart", "telegram-assistant"],
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
            )
            return "Перезапуск инициирован. Бот перезапустится через 2-3 секунды."
        elif name == "add_contact":
            return contacts_tool.add_contact(args["name"], args.get("phone",""), args.get("telegram",""), args.get("email",""), args.get("notes",""))
        elif name == "find_contact":
            return contacts_tool.find_contact(args["query"])
        elif name == "list_contacts":
            return contacts_tool.list_contacts()
        elif name == "delete_contact":
            return contacts_tool.delete_contact(args["contact_id"])
        elif name == "get_contact_card":
            return contacts_tool.format_contact_card(args["query"])
        elif name == "search_files":
            return email_tool.search_files(args["query"])
        elif name == "send_file_email":
            return email_tool.send_file_email(args["file_path"], args["to_name"], args.get("subject", ""))
        elif name == "check_send_status":
            return email_tool.check_send_status()
        elif name == "add_team_task":
            return team_tasks.add_team_task(
                args["executor"],
                args["task"],
                args.get("deadline", ""),
                args.get("result", ""),
                args.get("how_to_check", ""),
            )
        elif name == "save_card":
            card_type = args["card_type"]
            title = args["title"]
            content = args["content"]
            # Идеи → Google Sheets (таблица проектных заметок), остальное → vault
            if card_type == "idea":
                project_raw = args.get("project", "Без проекта") or "Без проекта"
                text_for_sheet = f"{title}: {content}" if title else content
                return _save_project_note_sync(project_raw, text_for_sheet)
            return vault_tool.save_card(card_type, title, content, args.get("tags", ""))
        elif name == "search_vault":
            return vault_tool.search_vault(args["query"])
        else:
            return f"Неизвестный инструмент: {name}"
    except Exception as e:
        return f"Ошибка инструмента {name}: {e}"

# ── Персистентный typing-индикатор ───────────────────────────
async def _keep_typing(chat_id: int, stop: asyncio.Event):
    while not stop.is_set():
        try:
            await bot.send_chat_action(chat_id, "typing")
        except Exception:
            pass
        try:
            await asyncio.wait_for(asyncio.shield(stop.wait()), timeout=4)
        except asyncio.TimeoutError:
            pass

# Инструменты с мгновенным системным подтверждением.
# Сюда НЕ входят add_reminder, add_note, cancel_reminder и т.п. —
# LLM сам пишет подтверждение в финальном ответе, дублировать не нужно.
_ACTION_TOOLS = {
    "send_file_email",  # высокорисковое действие — важно подтвердить сразу
}

# Состояние: user_id → fact_id (ожидаем исправление текста факта)
_awaiting_correction: dict[int, int] = {}
# Состояние: user_id → agreement_id (ожидаем исправление текста договорённости)
_pending_fix_state: dict[int, int] = {}


def _fact_approval_keyboard(fact_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="✅ Верно", callback_data=f"fact_ok_{fact_id}"),
        InlineKeyboardButton(text="✏️ Исправить", callback_data=f"fact_fix_{fact_id}"),
        InlineKeyboardButton(text="❌ Не то", callback_data=f"fact_no_{fact_id}"),
    ]])


async def _execute_and_notify(name: str, args: dict, user_id: int, chat_id: int) -> str:
    """Выполняет инструмент."""
    result = execute_tool(name, args, user_id)
    if name in _ACTION_TOOLS:
        await bot.send_message(chat_id, f"✅ {result}")
    return str(result)


# ── LLM-цикл ─────────────────────────────────────────────────
async def run_llm(history: list[dict], user_id: int, chat_id: int) -> str:
    recent_memory = mem_tool.get_recent_summary(5)
    memory_section = f"\n\n## Из памяти (последние факты):\n{recent_memory}" if recent_memory else ""

    knowledge_section = f"\n\n## База знаний (ГЛАВНЫЙ ИСТОЧНИК — читай в первую очередь):\n{KNOWLEDGE}" if KNOWLEDGE else ""

    system_msg = {
        "role": "system",
        "content": (
            f"Ты персональный ассистент Александра Бениаминова. "
            f"Сейчас {datetime.now(_MSK).strftime('%d.%m.%Y %H:%M')} МСК.\n\n"
            "⛔ ГЛАВНОЕ ПРАВИЛО — НИКОГДА НЕ НАРУШАТЬ:\n"
            "Вопросы о команде (Виктор, Евгения, Надежда, Управляющий, Тех.директор), "
            "проектах (Губаха, Монблан, Юникорн), KPI, задачах — отвечай ТОЛЬКО из базы знаний ниже. "
            "web_search для этого ЗАПРЕЩЁН. Не ищи этих людей в интернете — они сотрудники Александра.\n\n"
            f"{knowledge_section}"
            f"{memory_section}"
            "\n\n"
            "МОДЕЛЬ: Если задача требует глубокого анализа, составления длинного текста или "
            "сложного рассуждения — предложи: «Это лучше решит Claude Sonnet 4.6 — переключить?»\n\n"
            "СХЕМА РЕШЕНИЙ — выполняй первое подходящее:\n\n"
            "① Упомянут исполнитель из команды (Виктор / Евгения / Надежда / Управляющий / Тех.директор) + задача →\n"
            "   add_team_task. Если нет дедлайна — спроси: «Какой срок?». Как только срок есть — вызывай add_team_task немедленно.\n"
            "   ЗАПРЕЩЕНО использовать add_note или save_project_note для командных задач.\n"
            "   Фразы-триггеры: «задача Виктору», «скажи Надежде», «поставь Евгении», «добавь задачу», «запиши задачу» + имя члена команды.\n\n"
            "② Упомянуто время + действие → add_reminder.\n"
            "   Алиасы времени: «утро»=09:00, «день»=14:00, «вечер»=18:00,\n"
            "   «срочно»/«сейчас»=текущее время+2 минуты, «ночь»=22:00.\n"
            "   Сразу вызывай add_reminder — без уточнения «Верно?» и без эха.\n\n"
            "③ «Заметка для проекта X» / «Идея для X» → save_card(card_type='idea', title='', content=текст).\n"
            "   Из project_raw извлеки название проекта (Губаха/Монблан/Юникорн/Без проекта).\n"
            "   Если проект не назван — project_raw='Без проекта'.\n"
            "   Это пишет в Google Sheets — НЕ в add_note.\n\n"
            "③б Нет времени, нет проекта, личное дело/мысль → add_note (НЕ напоминание).\n\n"
            "④ «каждый день» / «каждую неделю» / «каждый месяц [число]» / «каждый год» →\n"
            "   add_reminder с параметром recurrence (daily/weekly/monthly/yearly).\n"
            "   remind_at = дата+время первого срабатывания в ISO8601.\n"
            "   После ✅ напоминание автоматически планируется на следующий период — Google Calendar НЕ нужен.\n"
            "   Уточни время если не указано. Не спрашивай «бессрочно?» — повторяется до ручной отмены.\n\n"
            "⑤ «что сегодня» / «что завтра» / «покажи [день]» / «расписание» → get_schedule(YYYY-MM-DD).\n\n"
            "⑥ «добавь встречу» / «добавь событие» → уточни дату и время → create_calendar_event.\n\n"
            "⑦ «запомни что...» / «запиши что...» / пользователь сообщает факт о человеке, проекте → remember_fact.\n"
            "   ВАЖНО: когда пользователь говорит «запомни что X» — вызови remember_fact(X) и БОЛЬШЕ НИЧЕГО.\n"
            "   Не вызывай recall_facts, web_search или другие инструменты — только remember_fact.\n"
            "   Если в одном сообщении несколько фактов — вызови remember_fact несколько раз.\n\n"
            "⑧ «какие заметки» / «какие идеи» / «что записано» → list_notes (из Google Sheets).\n\n"
            "⑨ Аудиофайл (.amr / .m4a / запись звонка) → transcribe → сформируй протокол:\n"
            "   👥 Участники: [стороны]\n"
            "   ✅ Договорённости: [список]\n"
            "   📋 Мои задачи: [задача — дедлайн если упоминался]\n"
            "   ❓ Открытые вопросы: [если есть, иначе пропусти]\n"
            "   После протокола — проанализируй дедлайны и договорённости, сам предложи:\n"
            "   «Добавить в Calendar: [событие] [предлагаемая дата]?»\n\n"
            "⑩ Вопрос о команде, проектах, KPI — ВСЕГДА отвечай из ## База знаний выше. Не говори «не знаю».\n\n"
            "⑪ Всё остальное → текстовый ответ.\n\n"
            "ЖЁСТКИЕ ЗАПРЕТЫ:\n"
            "- НЕЛЬЗЯ использовать web_search для вопросов о Викторе, Евгении, Надежде, Управляющем, Тех.директоре, Александре, его проектах и KPI.\n"
            "- НЕЛЬЗЯ устанавливать напоминание без явного времени — только заметка.\n"
            "- НЕЛЬЗЯ спрашивать «куда сохранить?» — решай сам по схеме выше.\n"
            "- НЕЛЬЗЯ угадывать время — нет времени = заметка.\n"
            "- НЕЛЬЗЯ дублировать напоминания другим людям — только личные.\n"
            "- НЕЛЬЗЯ перед add_reminder спрашивать «Верно?» или «Подтвердить?» — записывать сразу.\n"
            "- НЕЛЬЗЯ при «запомни что...» вызывать что-либо кроме remember_fact.\n"
            "- НЕЛЬЗЯ после add_team_task спрашивать «Поставить напоминание проверить выполнение?» — не задавать этот вопрос.\n\n"
            "СТИЛЬ: 1-2 предложения. После инструмента — одна строка подтверждения.\n"
            "ССЫЛКИ: Если использовал web_search — ВСЕГДА включай в ответ минимум 2-3 ссылки из результатов. "
            "Без ссылок отвечать ЗАПРЕЩЕНО. Формат: [Название](URL)."
        ),
    }
    try:
        sys.path.insert(0, "/home/parser/bots/shared")
        from rule_engine import get_system_addons
        addons = get_system_addons("assistant")
        if addons:
            system_msg["content"] += f"\n\nПРАВИЛА ПОЛЬЗОВАТЕЛЯ:\n{addons}"
    except Exception:
        pass

    messages = [system_msg] + history
    total_tool_calls = 0

    for _ in range(12):
        # Защита от бесконечного цикла: после 10 tool_calls — только текстовый ответ
        force_text = total_tool_calls >= 10
        create_kwargs: dict = {"model": config.MODEL, "messages": messages, "max_tokens": _MAX_TOKENS}
        if not force_text:
            create_kwargs["tools"] = TOOLS
            create_kwargs["tool_choice"] = "auto"
        resp = await ai_client.chat.completions.create(**create_kwargs)
        msg = resp.choices[0].message

        if not msg.tool_calls:
            content = msg.content or ""
            # DeepSeek иногда возвращает tool_calls в текстовом DSML-формате
            dsml_calls = _parse_dsml_tool_calls(content)
            if dsml_calls and total_tool_calls < 10:
                messages.append({"role": "assistant", "content": _strip_dsml(content) or ""})
                results_parts = []
                for name, args in dsml_calls:
                    result = await _execute_and_notify(name, args, user_id, chat_id)
                    results_parts.append(f"[{name}]: {result}")
                    total_tool_calls += 1
                messages.append({
                    "role": "user",
                    "content": "Результаты инструментов:\n\n" + "\n\n".join(results_parts) + "\n\nОтветь пользователю текстом.",
                })
                continue  # LLM получит результаты и ответит текстом или сделает ещё tool_calls
            return _strip_dsml(content) or "(пустой ответ)"

        # Выполняем все tool_calls этой итерации
        messages.append(msg.model_dump(exclude_unset=True))
        total_tool_calls += len(msg.tool_calls)

        for tc in msg.tool_calls:
            args = json.loads(tc.function.arguments)
            result = await _execute_and_notify(tc.function.name, args, user_id, chat_id)
            messages.append({
                "role": "tool",
                "tool_call_id": tc.id,
                "content": result,
            })
        # Продолжаем цикл — LLM сам решит: ещё tool_calls или финальный текстовый ответ.
        # Это позволяет выполнять многошаговые задачи (удалить 5 → создать 3 напоминания).

    return "Не удалось обработать запрос. Попробуй переформулировать."

# ── Транскрипция голоса ───────────────────────────────────────

# Известные галлюцинации Whisper — пустой результат или промпт вместо текста
_WHISPER_HALLUCINATIONS = {
    "transcribe the audio",
    "transcribe the video",
    "thank you for watching",
    "thanks for watching",
    "subtitle by",
    "subtitles by",
    "downloaded from",
    "you",
    "",
}

def _is_whisper_hallucination(text: str) -> bool:
    """Проверяет, является ли результат галлюцинацией Whisper."""
    cleaned = text.strip().lower().rstrip(".")
    if not cleaned:
        return True
    if cleaned in _WHISPER_HALLUCINATIONS:
        return True
    # Слишком короткий текст для 9+ секунд аудио (< 3 символов)
    if len(cleaned) < 3:
        return True
    return False

async def _transcribe_with_fallback(
    groq_client, audio_path: str, audio_name: str, mime: str, ru_prompt: str
) -> str | None:
    """Транскрибирует аудио: сначала turbo, при галлюцинации — large-v3.
    Возвращает текст или None если оба варианта дали галлюцинацию."""
    with open(audio_path, "rb") as f:
        tr = groq_client.audio.transcriptions.create(
            model="whisper-large-v3-turbo",
            file=(audio_name, f, mime),
            language="ru",
            prompt=ru_prompt,
        )
    result = tr.text.strip()

    # Если галлюцинация — повторяем через whisper-large-v3 (точнее, но медленнее)
    if _is_whisper_hallucination(result):
        logging.warning(f"[Whisper] turbo галлюцинация: {repr(result)}, пробуем whisper-large-v3")
        with open(audio_path, "rb") as f:
            tr2 = groq_client.audio.transcriptions.create(
                model="whisper-large-v3",
                file=(audio_name, f, mime),
                language="ru",
                prompt=ru_prompt,
            )
        result = tr2.text.strip()

    if _is_whisper_hallucination(result):
        return None
    return result


async def transcribe_voice(message: Message) -> str:
    from groq import Groq
    voice = message.voice
    file = await bot.get_file(voice.file_id)
    with tempfile.NamedTemporaryFile(suffix=".ogg", delete=False) as tmp:
        tmp_path = tmp.name
    try:
        await bot.download_file(file.file_path, tmp_path)

        # Проверяем что файл реально скачался
        file_size = os.path.getsize(tmp_path)
        if file_size < 100:
            return "[Не удалось распознать голос: пустой аудиофайл]"

        groq_client = Groq(api_key=config.GROQ_API_KEY)

        # Prompt подсказывает модели что это русская речь — снижает галлюцинации
        ru_prompt = "Это голосовое сообщение на русском языке."

        result = await _transcribe_with_fallback(groq_client, tmp_path, "voice.ogg", "audio/ogg", ru_prompt)

        # Если оба варианта дали галлюцинацию — сообщаем пользователю
        if result is None:
            logging.warning(f"[Whisper] оба варианта дали галлюцинацию, файл {file_size} байт")
            return "[Не удалось распознать голос: Whisper не смог разобрать аудио. Попробуй отправить ещё раз или напиши текстом.]"

        return result

    except Exception as e:
        logging.error(f"[Whisper] ошибка транскрипции: {e}")
        return f"[Не удалось распознать голос: {e}]"
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)

_WHISPER_SUPPORTED = {".mp3", ".mp4", ".m4a", ".wav", ".ogg", ".webm", ".flac", ".opus", ".mpeg", ".mpga"}
_NEEDS_CONVERT = {".amr", ".aac", ".3gp", ".3gpp", ".caf", ".wma"}

async def _convert_to_mp3(src_path: str) -> str:
    """Конвертирует аудио в mp3 через ffmpeg. Возвращает путь к новому файлу."""
    dst_path = src_path + ".mp3"
    proc = await asyncio.create_subprocess_exec(
        "ffmpeg", "-y", "-i", src_path, "-ar", "16000", "-ac", "1", "-b:a", "64k", dst_path,
        stdout=asyncio.subprocess.DEVNULL,
        stderr=asyncio.subprocess.DEVNULL,
    )
    await proc.wait()
    if not os.path.exists(dst_path) or os.path.getsize(dst_path) < 100:
        raise RuntimeError("ffmpeg не смог сконвертировать файл")
    return dst_path


async def transcribe_audio_file(file_id: str, filename: str = "audio.m4a") -> str:
    """Транскрибирует аудиофайл (запись звонка от Cube ACR и других источников)."""
    from groq import Groq
    ext = Path(filename).suffix.lower()
    tg_file = await bot.get_file(file_id)
    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
        tmp_path = tmp.name
    converted_path = None
    try:
        await bot.download_file(tg_file.file_path, tmp_path)
        if os.path.getsize(tmp_path) < 100:
            return "[Не удалось распознать: пустой аудиофайл]"

        # AMR и другие форматы не поддерживает Whisper — конвертируем через ffmpeg
        if ext in _NEEDS_CONVERT or ext not in _WHISPER_SUPPORTED:
            logging.info(f"[Whisper] конвертирую {ext} → mp3")
            converted_path = await _convert_to_mp3(tmp_path)
            audio_path, audio_name, mime = converted_path, "audio.mp3", "audio/mpeg"
        else:
            mime_map = {".mp3": "audio/mpeg", ".mp4": "audio/mp4", ".m4a": "audio/mp4",
                        ".wav": "audio/wav", ".ogg": "audio/ogg", ".webm": "audio/webm",
                        ".flac": "audio/flac", ".opus": "audio/opus"}
            audio_path, audio_name, mime = tmp_path, filename, mime_map.get(ext, "audio/mpeg")

        groq_client = Groq(api_key=config.GROQ_API_KEY)
        ru_prompt = "Это запись телефонного разговора на русском языке."
        result = await _transcribe_with_fallback(groq_client, audio_path, audio_name, mime, ru_prompt)
        if result is None:
            return "[Не удалось распознать запись звонка. Попробуй переслать файл ещё раз.]"
        return result
    except Exception as e:
        logging.error(f"[Whisper] ошибка транскрипции файла: {e}")
        return f"[Не удалось распознать: {e}]"
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
        if converted_path and os.path.exists(converted_path):
            os.unlink(converted_path)


_CALL_SUMMARY_SYSTEM = """Ты составляешь саммари телефонного разговора.
Выдели самое важное в структурированном виде. Будь конкретным — имена, цифры, даты.
Если чего-то нет в разговоре — не придумывай, просто пропусти блок."""

async def _generate_call_summary(transcript: str, user_id: int, chat_id: int) -> str:
    prompt = f"""Составь саммари этого телефонного разговора:

{transcript}

Формат:
👥 *Участники:* [имена или описания сторон]

✅ *Договорённости:*
• [каждая договорённость отдельной строкой]

📋 *Мои задачи:*
• [задача — дедлайн если упоминался]

❓ *Открытые вопросы:* [если есть, иначе пропусти блок]"""
    try:
        resp = await ai_client.chat.completions.create(
            model=config.MODEL,
            messages=[
                {"role": "system", "content": _CALL_SUMMARY_SYSTEM},
                {"role": "user", "content": prompt},
            ],
            max_tokens=600,
            temperature=0.3,
        )
        return resp.choices[0].message.content.strip()
    except Exception as e:
        logging.error(f"[CallSummary] ошибка: {e}")
        return "Не удалось составить саммари. Транскрипция выше — используй её вручную."


async def _extract_and_save_agreements(summary: str) -> int:
    """Извлекает конкретные задачи и договорённости из саммари → сохраняет в agreements. Возвращает кол-во."""
    from tools.db import add_agreement
    extract_prompt = f"""Из этого резюме звонка извлеки список конкретных задач и договорённостей.
Верни JSON-массив: [{{"text": "..."}}]
Только конкретные действия (позвонить, отправить, встретиться, договорились и т.д.).
Не включай общие факты и описания. Если нечего извлечь — верни [].

Резюме:
{summary}"""
    try:
        resp = await ai_client.chat.completions.create(
            model=config.MODEL,
            messages=[
                {"role": "system", "content": "Извлеки задачи и договорённости из резюме. Верни только JSON-массив."},
                {"role": "user", "content": extract_prompt},
            ],
            max_tokens=400,
            temperature=0.1,
        )
        raw = resp.choices[0].message.content.strip()
        raw = re.sub(r'```json\s*', '', raw)
        raw = re.sub(r'```\s*', '', raw)
        items = json.loads(raw)
        count = 0
        for item in items:
            if isinstance(item, dict) and item.get("text"):
                add_agreement(item["text"].strip(), source="call")
                count += 1
        return count
    except Exception as e:
        logging.warning(f"[Agreements] не удалось извлечь: {e}")
        return 0


def _agreement_kb(agreement_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="📅 В календарь", callback_data=f"agr_cal_{agreement_id}"),
        InlineKeyboardButton(text="✏️ Исправить", callback_data=f"agr_fix_{agreement_id}"),
        InlineKeyboardButton(text="🗑 Удалить", callback_data=f"agr_del_{agreement_id}"),
    ]])


# ── Доступ ────────────────────────────────────────────────────
def is_allowed(user_id: int) -> bool:
    return not config.ALLOWED_USER_IDS or user_id in config.ALLOWED_USER_IDS

# ── Хендлеры ─────────────────────────────────────────────────
_HELP_TEXT = """🤖 *Напоминатор — шпаргалка*

*⏰ Напоминания*
`Напомни завтра в 9:30 позвонить Виктору`
`Напомни каждый понедельник в 10:00 сделать отчёт`
`Напомни через 2 часа проверить духовку`
Повторяет каждые 30 мин пока не нажмёшь ✅

*📅 Календарь*
`Что у меня сегодня / завтра?`
`Создай встречу "Планёрка" в среду в 11:00`

*💡 Идеи → Google Sheets*
`Сохрани идею: запустить завтрак-буфет по выходным`
`Запомни идею про маркетинг Губахи`
Идеи автоматически попадают в таблицу идей

*🧠 Решения и выводы → база знаний*
`Запомни решение: партнёрство с туроператором — нет`
`Зафиксируй вывод: гости бронируют поздно, нужен триггер за 2 дня`
`Найди в базе знаний решения по Губахе`

*💾 Память фактов*
`Запомни: Виктор предпочитает задачи письменно`
`Что ты помнишь про Евгению?`

*📝 Заметки*
`Сохрани заметку: обновить меню до 15-го`
`Покажи последние заметки` → /notes

*🌐 Веб-поиск*
`Найди актуальный курс доллара`
`Что пишут про горнолыжный сезон на Урале?`

*👥 Задачи команде Губаха*
`Поставь задачу Надежде: пост про майские акции, срок пятница`
`Скажи Евгении: проверить загрузку на TravelLine, срок завтра`

*📁 Файлы на Mac*
`Найди файл про смету Губахи`
`Отправь договор_монблан.pdf Кате`

*🎙 Голос* — просто говори, бот расшифрует и выполнит
*📎 Документ/фото* — прикрепи, бот прочтёт и ответит на вопросы
*📞 Аудио звонка* — пришли запись, получишь структурированное резюме

*Команды:*
/help — эта шпаргалка
/notes — последние заметки
/reminders — активные напоминания
/memory — сохранённые факты
/reset — очистить контекст диалога"""


@dp.message(Command("start"))
async def cmd_start(message: Message):
    if not is_allowed(message.from_user.id):
        return
    await message.answer(
        "Привет! Я персональный ассистент Александра.\n\n"
        "Напиши /help — покажу все команды и примеры.",
    )

@dp.message(Command("whoami"))
async def cmd_whoami(message: Message):
    uid = message.from_user.id
    await message.answer(
        f"Telegram ID: `{uid}`\nИмя: {message.from_user.full_name}",
        parse_mode="Markdown",
    )

@dp.message(Command("reset"))
async def cmd_reset(message: Message):
    if not is_allowed(message.from_user.id):
        return
    histories.pop(message.from_user.id, None)
    await message.answer("История диалога очищена.")

@dp.message(Command("notes"))
async def cmd_notes(message: Message):
    if not is_allowed(message.from_user.id):
        return
    await message.answer(notes.list_notes(10))

@dp.message(Command("reminders"))
async def cmd_reminders(message: Message):
    if not is_allowed(message.from_user.id):
        return
    await message.answer(rem_tool.list_reminders(message.from_user.id))

@dp.message(Command("memory"))
async def cmd_memory(message: Message):
    if not is_allowed(message.from_user.id):
        return
    await message.answer(mem_tool.list_memories())


@dp.message(Command("facts"))
async def cmd_facts(message: Message):
    """Показывает факты, ожидающие подтверждения."""
    if not is_allowed(message.from_user.id):
        return
    pending = mem_tool.get_pending_facts()
    if not pending:
        await message.answer("✅ Нет фактов на проверке.")
        return
    await message.answer(f"⏳ Факты на проверке ({len(pending)}):")
    for f in pending:
        await message.answer(
            f"#{f['id']} [{f['category']}]: {f['text']}",
            reply_markup=_fact_approval_keyboard(f["id"]),
        )

@dp.message(Command("help"))
async def cmd_help(message: Message):
    if not is_allowed(message.from_user.id):
        return
    await message.answer(_HELP_TEXT, parse_mode="Markdown")

@dp.message(Command("restart"))
async def cmd_restart(message: Message):
    if not is_allowed(message.from_user.id):
        return
    await message.answer("♻️ Перезапускаю бота...")
    import subprocess
    subprocess.Popen(
        ["systemctl", "restart", "telegram-assistant"],
        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
    )

# ── Кнопки напоминаний ────────────────────────────────────────
@dp.callback_query(F.data.startswith("ack_"))
async def handle_ack(callback: CallbackQuery):
    if not is_allowed(callback.from_user.id):
        return
    reminder_id = int(callback.data.split("_")[1])
    rem_tool.ack_reminder_by_id(reminder_id, callback.from_user.id)
    await callback.message.edit_text(
        callback.message.text + "\n\n✅ _Выполнено_",
        parse_mode="Markdown",
        reply_markup=None,
    )
    await callback.answer("Выполнено!")

@dp.callback_query(F.data.startswith("snz_"))
async def handle_snooze(callback: CallbackQuery):
    if not is_allowed(callback.from_user.id):
        return
    parts = callback.data.split("_")
    reminder_id = int(parts[1])

    if len(parts) == 2:
        await callback.message.edit_reply_markup(
            reply_markup=rem_tool._snooze_keyboard(reminder_id)
        )
        await callback.answer()
    elif parts[2] == "back":
        await callback.message.edit_reply_markup(
            reply_markup=rem_tool._main_keyboard(reminder_id)
        )
        await callback.answer()
    else:
        minutes = 0 if parts[2] == "tmr" else int(parts[2])
        label = rem_tool.snooze_reminder(reminder_id, callback.from_user.id, minutes)
        await callback.message.edit_text(
            callback.message.text + f"\n\n⏰ _Отложено: {label}_",
            parse_mode="Markdown",
            reply_markup=None,
        )
        await callback.answer(f"Отложено: {label}")

# ── Подтверждение фактов ─────────────────────────────────────
@dp.callback_query(F.data.startswith("fact_ok_"))
async def handle_fact_approve(callback: CallbackQuery):
    if not is_allowed(callback.from_user.id):
        return
    fact_id = int(callback.data.split("_")[2])
    result = mem_tool.approve_fact(fact_id)
    await callback.message.edit_text(f"✅ {result}", reply_markup=None)
    await callback.answer("Сохранено!")


@dp.callback_query(F.data.startswith("fact_fix_"))
async def handle_fact_fix(callback: CallbackQuery):
    if not is_allowed(callback.from_user.id):
        return
    fact_id = int(callback.data.split("_")[2])
    _awaiting_correction[callback.from_user.id] = fact_id
    await callback.message.edit_text(
        callback.message.text + "\n\n✏️ _Введи правильный текст:_",
        parse_mode="Markdown",
        reply_markup=None,
    )
    await callback.answer()


@dp.callback_query(F.data.startswith("fact_no_"))
async def handle_fact_deny(callback: CallbackQuery):
    if not is_allowed(callback.from_user.id):
        return
    fact_id = int(callback.data.split("_")[2])
    mem_tool.reject_fact(fact_id)
    await callback.message.edit_text("❌ Факт отклонён", reply_markup=None)
    await callback.answer("Отклонено")


# ── Анализ документов (PDF, Word, Excel, изображения) ────────

_MAX_DOC_SIZE = 20 * 1024 * 1024  # 20 МБ — лимит Telegram Bot API


async def _download_doc(doc) -> bytes | None:
    if doc.file_size and doc.file_size > _MAX_DOC_SIZE:
        return None
    tg_file = await bot.get_file(doc.file_id)
    buf = await bot.download_file(tg_file.file_path)
    return buf.read()


async def _analyze_pdf(raw: bytes, hint: str) -> str:
    import pdfplumber, io as _io
    text = ""
    with pdfplumber.open(_io.BytesIO(raw)) as pdf:
        for page in pdf.pages[:15]:
            text += (page.extract_text() or "") + "\n"
    text = text.strip()

    if len(text) > 200:
        prompt = hint or "Проанализируй документ. Выдели: тип, стороны, ключевые условия, сроки, суммы."
        resp = await ai_client.chat.completions.create(
            model=config.MODEL,
            messages=[
                {"role": "system", "content": "Анализируй юридический/деловой документ. Отвечай структурированно по-русски."},
                {"role": "user", "content": f"{prompt}\n\n---\n{text[:12000]}"},
            ],
            max_tokens=_MAX_TOKENS,
        )
        return (resp.choices[0].message.content or "").strip()

    # Скан → конвертируем в картинки и отправляем в vision
    import tempfile, os, base64, io as _io2
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        tmp.write(raw); tmp_path = tmp.name
    try:
        from pdf2image import convert_from_path
        images = convert_from_path(tmp_path, dpi=150, first_page=1, last_page=3)
        if not images:
            return "Не удалось извлечь текст из PDF."
        content: list = []
        for img in images[:2]:
            buf = _io2.BytesIO()
            img.save(buf, format="JPEG", quality=85)
            content.append({"type": "image_url", "image_url": {
                "url": f"data:image/jpeg;base64,{base64.b64encode(buf.getvalue()).decode()}"
            }})
        content.append({"type": "text", "text": hint or "Проанализируй документ. Выдели: тип, стороны, ключевые условия, сроки, суммы."})
        resp = await ai_client.chat.completions.create(
            model=config.MODEL,
            messages=[{"role": "user", "content": content}],
            max_tokens=_MAX_TOKENS,
        )
        return (resp.choices[0].message.content or "").strip()
    finally:
        os.unlink(tmp_path)


async def _analyze_docx(raw: bytes, hint: str) -> str:
    import docx, io as _io
    doc = docx.Document(_io.BytesIO(raw))
    text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    if not text:
        return "Документ Word пустой или не содержит текста."
    prompt = hint or "Проанализируй документ. Выдели: тип, стороны, ключевые условия, сроки, суммы."
    resp = await ai_client.chat.completions.create(
        model=config.MODEL,
        messages=[
            {"role": "system", "content": "Анализируй документ. Отвечай структурированно по-русски."},
            {"role": "user", "content": f"{prompt}\n\n---\n{text[:12000]}"},
        ],
        max_tokens=_MAX_TOKENS,
    )
    return (resp.choices[0].message.content or "").strip()


async def _analyze_xlsx(raw: bytes, hint: str) -> str:
    import openpyxl, io as _io
    wb = openpyxl.load_workbook(_io.BytesIO(raw), read_only=True, data_only=True)
    parts = []
    for name in wb.sheetnames[:3]:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(max_row=60, values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(c.strip() for c in cells):
                rows.append(" | ".join(cells))
        if rows:
            parts.append(f"Лист «{name}»:\n" + "\n".join(rows[:60]))
    text = "\n\n".join(parts)
    if not text:
        return "Excel-файл пустой или не содержит данных."
    prompt = hint or "Проанализируй таблицу. Что здесь за данные? Выдели ключевые показатели."
    resp = await ai_client.chat.completions.create(
        model=config.MODEL,
        messages=[
            {"role": "system", "content": "Анализируй таблицу. Отвечай по-русски."},
            {"role": "user", "content": f"{prompt}\n\n---\n{text[:8000]}"},
        ],
        max_tokens=1500,
    )
    return (resp.choices[0].message.content or "").strip()


async def _analyze_image_doc(raw: bytes, mime: str, hint: str) -> str:
    import base64
    content = [
        {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{base64.b64encode(raw).decode()}"}},
        {"type": "text", "text": hint or "Проанализируй изображение. Опиши что видишь."},
    ]
    resp = await ai_client.chat.completions.create(
        model=config.MODEL,
        messages=[{"role": "user", "content": content}],
        max_tokens=1500,
    )
    return (resp.choices[0].message.content or "").strip()


async def _handle_document_file(message: Message):
    """Роутер для не-аудио документов: PDF, Word, Excel, изображения."""
    user_id = message.from_user.id
    doc = message.document
    filename = doc.file_name or "document"
    ext = Path(filename).suffix.lower()
    mime = doc.mime_type or ""
    hint = (message.caption or "").strip()

    mb = round((doc.file_size or 0) / 1024 / 1024, 1)
    if doc.file_size and doc.file_size > _MAX_DOC_SIZE:
        await message.answer(
            f"⚠️ Файл слишком большой ({mb} МБ). "
            f"Telegram позволяет скачивать файлы до 20 МБ.\n"
            f"Попробуй сжать PDF или разбить на части."
        )
        return

    stop = asyncio.Event()
    typing = asyncio.create_task(_keep_typing(message.chat.id, stop))
    try:
        await message.answer(f"📄 Читаю «{filename}»…")
        raw = await _download_doc(doc)
        if raw is None:
            await message.answer("Не удалось скачать файл.")
            return

        if ext == ".pdf" or "pdf" in mime:
            result = await _analyze_pdf(raw, hint)
        elif ext in (".docx", ".doc") or "word" in mime or "officedocument.word" in mime:
            result = await _analyze_docx(raw, hint)
        elif ext in (".xlsx", ".xls") or "excel" in mime or "spreadsheet" in mime:
            result = await _analyze_xlsx(raw, hint)
        elif ext == ".csv":
            text = raw.decode("utf-8", errors="replace")[:8000]
            prompt = hint or "Проанализируй CSV-данные. Выдели ключевые показатели."
            resp = await ai_client.chat.completions.create(
                model=config.MODEL,
                messages=[{"role": "user", "content": f"{prompt}\n\n{text}"}],
                max_tokens=1500,
            )
            result = (resp.choices[0].message.content or "").strip()
        elif mime.startswith("image/") or ext in (".jpg", ".jpeg", ".png", ".webp"):
            result = await _analyze_image_doc(raw, mime or "image/jpeg", hint)
        else:
            await message.answer(
                f"❓ Формат «{ext or mime}» не поддерживается.\n"
                f"Поддерживаю: PDF, Word (.docx), Excel (.xlsx), CSV, JPEG/PNG."
            )
            return
    except Exception as e:
        logging.error(f"[doc] ошибка анализа {filename}: {e}")
        result = f"Ошибка при анализе файла: {e}"
    finally:
        stop.set(); typing.cancel()
        try: await typing
        except asyncio.CancelledError: pass

    # Добавляем в историю чтобы можно было задавать уточняющие вопросы
    history = histories.setdefault(user_id, [])
    history.append({"role": "user", "content": f"[Файл: {filename}]{chr(10) + hint if hint else ''}"})
    history.append({"role": "assistant", "content": result})
    _last_message_time[user_id] = datetime.now().timestamp()

    for chunk in range(0, len(result), 4000):
        await message.answer(result[chunk:chunk+4000])


# ── Запись звонков → саммари ─────────────────────────────────
@dp.message(F.audio | F.document)
async def handle_audio_file(message: Message):
    user_id = message.from_user.id
    if not is_allowed(user_id):
        return
    if message.document:
        mime = message.document.mime_type or ""
        ext = Path(message.document.file_name or "").suffix.lower()
        # Не-аудио документы → отдельный обработчик
        if not (mime.startswith("audio/") or ext in _WHISPER_SUPPORTED or ext in _NEEDS_CONVERT):
            await _handle_document_file(message)
            return
        file_id = message.document.file_id
        filename = message.document.file_name or "audio.m4a"
    else:
        file_id = message.audio.file_id
        filename = message.audio.file_name or "audio.m4a"

    await message.answer("🎙 Транскрибирую запись звонка...")
    transcript = await transcribe_audio_file(file_id, filename)
    if transcript.startswith("[Не удалось"):
        await message.answer(transcript)
        return

    # Транскрипция — отправляем отдельно (может быть длинной)
    preview = transcript if len(transcript) <= 3000 else transcript[:3000] + "…"
    await message.answer(f"📝 *Транскрипция:*\n_{preview}_", parse_mode="Markdown")

    await message.answer("🤔 Составляю саммари...")
    summary = await _generate_call_summary(transcript, user_id, message.chat.id)
    _pending_call_summaries[user_id] = summary
    _pending_call_transcripts[user_id] = transcript

    kb = _call_summary_kb()
    await message.answer(f"📋 *Саммари звонка:*\n\n{summary}", reply_markup=kb, parse_mode="Markdown")

    # Извлекаем договорённости в фоне — не задерживаем ответ пользователю
    asyncio.create_task(_extract_and_save_agreements(summary))


def _call_summary_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="📅 Напоминания", callback_data="call_remind"),
            InlineKeyboardButton(text="💾 В заметки", callback_data="call_save"),
        ],
        [
            InlineKeyboardButton(text="✏️ Исправить", callback_data="call_edit"),
            InlineKeyboardButton(text="🗑 Удалить", callback_data="call_delete"),
        ],
    ])


@dp.callback_query(F.data == "call_remind")
async def call_remind_callback(callback: CallbackQuery):
    user_id = callback.from_user.id
    summary = _pending_call_summaries.get(user_id, "")
    if not summary:
        await callback.answer("Саммари не найдено")
        return
    await callback.message.edit_reply_markup(reply_markup=None)
    # Передаём саммари в основной обработчик как запрос на создание напоминаний
    history = histories.setdefault(user_id, [])
    history.append({"role": "user", "content": f"Создай напоминания по задачам из этого саммари звонка:\n{summary}"})
    response = await run_llm(history, user_id, callback.message.chat.id)
    history.append({"role": "assistant", "content": response})
    await callback.message.answer(response)
    await callback.answer()


@dp.callback_query(F.data == "call_save")
async def call_save_callback(callback: CallbackQuery):
    user_id = callback.from_user.id
    summary = _pending_call_summaries.pop(user_id, "")
    if not summary:
        await callback.answer("Саммари не найдено")
        return
    await callback.message.edit_reply_markup(reply_markup=None)
    note_text = f"Звонок {datetime.now(_MSK).strftime('%d.%m.%Y %H:%M')}:\n{summary}"
    result = notes.add_note(note_text)
    await callback.message.answer(f"💾 {result}")
    await callback.answer()


@dp.callback_query(F.data == "call_edit")
async def call_edit_callback(callback: CallbackQuery):
    user_id = callback.from_user.id
    if user_id not in _pending_call_summaries:
        await callback.answer("Саммари не найдено")
        return
    _awaiting_call_correction[user_id] = callback.message.message_id
    await callback.message.edit_reply_markup(reply_markup=None)
    await callback.message.answer("✏️ Напиши что нужно исправить или уточнить в саммари:")
    await callback.answer()


@dp.callback_query(F.data == "call_delete")
async def call_delete_callback(callback: CallbackQuery):
    user_id = callback.from_user.id
    _pending_call_summaries.pop(user_id, None)
    _pending_call_transcripts.pop(user_id, None)
    _awaiting_call_correction.pop(user_id, None)
    await callback.message.edit_text("🗑 Саммари удалено.")
    await callback.answer()


# ── Дайджест договорённостей — кнопки ────────────────────────

@dp.callback_query(F.data.startswith("agr_del_"))
async def agr_del_callback(callback: CallbackQuery):
    agr_id = int(callback.data.split("agr_del_")[1])
    from tools.db import update_agreement_status
    update_agreement_status(agr_id, "deleted")
    await callback.message.edit_text("🗑 Удалено.")
    await callback.answer()


@dp.callback_query(F.data.startswith("agr_fix_"))
async def agr_fix_callback(callback: CallbackQuery):
    agr_id = int(callback.data.split("agr_fix_")[1])
    user_id = callback.from_user.id
    _pending_fix_state[user_id] = agr_id
    await callback.message.edit_reply_markup(reply_markup=None)
    await callback.message.answer("✏️ Напиши исправленный текст:")
    await callback.answer()


@dp.callback_query(F.data.startswith("agr_cal_"))
async def agr_cal_callback(callback: CallbackQuery):
    from tools.db import get_agreement_by_id, update_agreement_status
    agr_id = int(callback.data.split("agr_cal_")[1])
    agr = get_agreement_by_id(agr_id)
    if not agr:
        await callback.answer("Не найдено")
        return
    await callback.message.edit_reply_markup(reply_markup=None)
    await callback.message.answer("📅 Добавляю в Google Calendar...")
    parse_prompt = (
        f"Из текста извлеки дату, время и название для события Google Calendar.\n"
        f"Текст: {agr['text']}\n"
        f"Сегодня: {datetime.now(_MSK).strftime('%d.%m.%Y')}\n"
        f"Верни JSON: {{\"date\": \"YYYY-MM-DD\", \"time\": \"HH:MM\", \"title\": \"название\"}}\n"
        f"Если дата не указана — null. Если время не указано — null."
    )
    try:
        resp = await ai_client.chat.completions.create(
            model=config.MODEL,
            messages=[
                {"role": "system", "content": "Извлеки дату и название события. Верни только JSON."},
                {"role": "user", "content": parse_prompt},
            ],
            max_tokens=150,
            temperature=0.0,
        )
        raw = resp.choices[0].message.content.strip()
        raw = re.sub(r'```json\s*', '', raw)
        raw = re.sub(r'```\s*', '', raw)
        parsed = json.loads(raw)
    except Exception as e:
        await callback.message.answer(f"Не удалось распознать дату: {e}\nДобавь вручную в календарь.")
        await callback.answer()
        return
    if not parsed.get("date"):
        await callback.message.answer("Не нашёл дату в тексте. Уточни когда это должно быть:")
        await callback.answer()
        return
    date_str = parsed["date"]
    time_str = parsed.get("time") or "10:00"
    start_iso = f"{date_str}T{time_str}:00"
    title = (parsed.get("title") or agr["text"])[:80]
    try:
        loop = asyncio.get_running_loop()
        result_msg = await loop.run_in_executor(
            None,
            lambda: calendar.create_event(
                config.GOOGLE_CALENDAR_ID,
                config.SERVICE_ACCOUNT_JSON,
                title,
                start_iso,
            )
        )
        update_agreement_status(agr_id, "calendar")
        await callback.message.answer(f"✅ {result_msg}", parse_mode="Markdown")
    except Exception as e:
        await callback.message.answer(f"Ошибка создания события: {e}")
    await callback.answer()


# ── Управление правилами прямо в чате ────────────────────────

def _shared_db():
    if "/home/parser/bots/shared" not in sys.path:
        sys.path.insert(0, "/home/parser/bots/shared")
    import rules_db
    return rules_db

def _shared_engine():
    if "/home/parser/bots/shared" not in sys.path:
        sys.path.insert(0, "/home/parser/bots/shared")
    import rule_engine
    return rule_engine

_BOT_NAME = "assistant"

RULE_TYPE_NAMES = {
    "system_addon": "📌 В промпт",
    "reformat": "✏️ Переформат",
    "append": "➕ Дописать в конец",
    "prepend": "⬆️ Вставить в начало",
}


async def _do_reformat(message: Message, text: str) -> None:
    user_id = message.from_user.id
    instruction = re.sub(r'^переделай\s*[—–-]?\s*', '', text, flags=re.IGNORECASE).strip()
    if not instruction:
        await message.answer("Укажи инструкцию: «переделай — пиши короче»")
        return
    original = (message.reply_to_message.text or "").strip()
    if not original:
        await message.answer("Не удалось получить текст исходного сообщения.")
        return
    await message.answer("✏️ Переформатирую...")
    try:
        resp = await ai_client.chat.completions.create(
            model=config.MODEL,
            messages=[
                {"role": "system", "content": "Переформатируй текст согласно инструкции. Верни только результат."},
                {"role": "user", "content": f"Инструкция: {instruction}\n\nТекст:\n{original}"},
            ],
            max_tokens=_MAX_TOKENS,
        )
        reformatted = (resp.choices[0].message.content or original).strip()
    except Exception as e:
        await message.answer(f"Ошибка: {e}")
        return
    _pending_reformat[user_id] = {"instruction": instruction, "original": original}
    kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="💾 Сохранить как правило", callback_data=f"save_rule:{_BOT_NAME}"),
        InlineKeyboardButton(text="✅ Готово", callback_data="reformat_done"),
    ]])
    await message.answer(f"*Результат:*\n\n{reformatted}\n\n_Сохранить это как правило для следующих ответов?_",
                         parse_mode="Markdown", reply_markup=kb)


@dp.callback_query(F.data.startswith("save_rule:"))
async def cb_save_rule(callback: CallbackQuery):
    bot_name = callback.data.split(":")[1]
    pending = _pending_reformat.pop(callback.from_user.id, None)
    if not pending:
        await callback.answer("Данные истекли.")
        return
    try:
        db = _shared_db()
        rule_id = db.create_rule(bot_name, "reformat", pending["instruction"],
                                 description=f"Авто: {pending['instruction'][:60]}")
        _shared_engine().invalidate_cache(bot_name)
        await callback.message.edit_text(
            f"💾 Правило #{rule_id} сохранено.\nБуду так форматировать следующие ответы."
        )
    except Exception as e:
        await callback.message.edit_text(f"Ошибка сохранения: {e}")
    await callback.answer()


@dp.callback_query(F.data == "reformat_done")
async def cb_reformat_done(callback: CallbackQuery):
    _pending_reformat.pop(callback.from_user.id, None)
    await callback.message.edit_reply_markup(reply_markup=None)
    await callback.answer("Готово.")


@dp.callback_query(F.data.startswith("del_rule:"))
async def cb_del_rule_confirm(callback: CallbackQuery):
    rule_id = int(callback.data.split(":")[1])
    try:
        db = _shared_db()
        ok = db.delete_rule(rule_id)
        _shared_engine().invalidate_cache()
        text = f"🗑 Правило #{rule_id} удалено." if ok else f"Правило #{rule_id} не найдено."
    except Exception as e:
        text = f"Ошибка: {e}"
    await callback.message.edit_text(text)
    await callback.answer()


@dp.callback_query(F.data == "del_cancel")
async def cb_del_cancel(callback: CallbackQuery):
    await callback.message.edit_text("Отменено.")
    await callback.answer()


@dp.message(Command("rules"))
async def cmd_rules(message: Message):
    if not is_allowed(message.from_user.id):
        return
    try:
        rules = _shared_db().list_rules(_BOT_NAME)
    except Exception as e:
        await message.answer(f"Ошибка: {e}")
        return
    if not rules:
        await message.answer("Правил нет. Процитируй мой ответ и напиши «переделай — инструкция».")
        return
    lines = []
    for r in rules:
        status = "✅" if r["active"] else "❌"
        type_label = RULE_TYPE_NAMES.get(r["rule_type"], r["rule_type"])
        desc = r["description"] or r["instruction"][:60]
        lines.append(f"{status} #{r['id']} {type_label}\n    {desc}")
    await message.answer("📋 *Мои правила:*\n\n" + "\n\n".join(lines), parse_mode="Markdown")


@dp.message(Command("toggle_rule"))
async def cmd_toggle_rule(message: Message):
    if not is_allowed(message.from_user.id):
        return
    parts = (message.text or "").split()
    if len(parts) < 2 or not parts[1].isdigit():
        await message.answer("Использование: /toggle_rule <id>")
        return
    rule_id = int(parts[1])
    try:
        new_state = _shared_db().toggle_rule(rule_id)
        _shared_engine().invalidate_cache()
    except Exception as e:
        await message.answer(f"Ошибка: {e}")
        return
    if new_state is None:
        await message.answer(f"Правило #{rule_id} не найдено.")
        return
    emoji = "✅" if new_state else "❌"
    await message.answer(f"{emoji} Правило #{rule_id} {'активно' if new_state else 'выключено'}.")


@dp.message(Command("delete_rule"))
async def cmd_delete_rule(message: Message):
    if not is_allowed(message.from_user.id):
        return
    parts = (message.text or "").split()
    if len(parts) < 2 or not parts[1].isdigit():
        await message.answer("Использование: /delete_rule <id>")
        return
    rule_id = int(parts[1])
    kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="✅ Удалить", callback_data=f"del_rule:{rule_id}"),
        InlineKeyboardButton(text="❌ Отмена", callback_data="del_cancel"),
    ]])
    await message.answer(f"Удалить правило #{rule_id}?", reply_markup=kb)


# ── Обработка сообщений ───────────────────────────────────────
@dp.message(F.text | F.voice)
async def handle_message(message: Message):
    user_id = message.from_user.id
    if not is_allowed(user_id):
        return

    # Сброс контекста при паузе > 5 минут
    now_ts = datetime.now().timestamp()
    prev_ts = _last_message_time.get(user_id, 0)
    _last_message_time[user_id] = now_ts
    if prev_ts > 0 and (now_ts - prev_ts) > _CONTEXT_TTL:
        histories.pop(user_id, None)

    # Режим исправления саммари звонка
    if user_id in _awaiting_call_correction:
        _awaiting_call_correction.pop(user_id)
        transcript = _pending_call_transcripts.get(user_id, "")
        old_summary = _pending_call_summaries.get(user_id, "")
        correction = message.text if message.text else (await transcribe_voice(message) if message.voice else "")
        if not correction or correction.startswith("[Не удалось"):
            await message.answer("Не удалось распознать корректировку. Попробуй ещё раз.")
            return
        await message.answer("🤔 Обновляю саммари...")
        prompt = f"""Исходная транскрипция звонка:\n{transcript}\n\nПредыдущее саммари:\n{old_summary}\n\nКорректировка от пользователя: {correction}\n\nСоставь обновлённое саммари с учётом корректировки, в том же формате."""
        try:
            resp = await ai_client.chat.completions.create(
                model=config.MODEL,
                messages=[
                    {"role": "system", "content": _CALL_SUMMARY_SYSTEM},
                    {"role": "user", "content": prompt},
                ],
                max_tokens=600,
                temperature=0.3,
            )
            new_summary = resp.choices[0].message.content.strip()
        except Exception as e:
            await message.answer(f"Ошибка обновления: {e}")
            return
        _pending_call_summaries[user_id] = new_summary
        await message.answer(f"📋 *Обновлённое саммари:*\n\n{new_summary}", reply_markup=_call_summary_kb(), parse_mode="Markdown")
        return

    # Режим исправления договорённости из дайджеста
    if message.text and user_id in _pending_fix_state:
        agr_id = _pending_fix_state.pop(user_id)
        from tools.db import update_agreement_status
        update_agreement_status(agr_id, "pending", text=message.text.strip())
        await message.answer("✅ Текст обновлён.")
        return

    # Режим исправления факта: следующее сообщение — новый текст факта
    if message.text and user_id in _awaiting_correction:
        fact_id = _awaiting_correction.pop(user_id)
        result = mem_tool.correct_fact(fact_id, message.text.strip())
        await message.answer(f"✅ {result}")
        return

    if message.voice:
        text = await transcribe_voice(message)
        if text.startswith("[Не удалось"):
            await message.answer(text)
            return
        await message.answer(f"🎙 _{text}_", parse_mode="Markdown")
    else:
        text = message.text

    # Пишем в daily vault для ночной обработки
    try:
        label = "🎙" if message.voice else "💬"
        vault_tool.append_daily(f"{label} {text}")
    except Exception:
        pass

    # Перехват «переделай» — reply на мой ответ
    if (message.reply_to_message
            and message.reply_to_message.from_user
            and message.reply_to_message.from_user.id == bot.id
            and text.lower().startswith("переделай")):
        await _do_reformat(message, text)
        return

    # ── Детектор «Заметка/Идея [для проекта X]» → запись в Google Sheets ──
    _idea_match = re.match(
        r"(?:заметка|идея)\s+(?:для\s+)?(?:проекта?\s+)?"
        r"(губах[аиу]?|юникорн|unicorn|монблан[аe]?|прочее?|офис|разное|общее|без\s+проекта)?"
        r"[\s:—\-]*(.+)",
        text.strip(), re.IGNORECASE | re.DOTALL
    )
    if _idea_match:
        _proj_raw = (_idea_match.group(1) or "Без проекта").strip()
        _proj_text = _idea_match.group(2).strip()
        if _proj_text:
            await message.answer(await _save_project_note(_proj_raw, _proj_text), parse_mode="Markdown")
            return
    # ── Детектор «проект X» → запись в Google Sheets ──────────────
    _project_match = re.match(
        r"(?:для\s+)?(?:проект(?:а|е|у)?\s+)?(губах[аиу]?|юникорн|unicorn|монблан[аe]?|прочее?|офис|разное|общее|без\s+проекта)\W+(.+)",
        text.strip(), re.IGNORECASE | re.DOTALL
    )
    if _project_match:
        _proj_raw = _project_match.group(1).strip()
        _proj_text = _project_match.group(2).strip()
        await message.answer(await _save_project_note(_proj_raw, _proj_text), parse_mode="Markdown")
        return
    # ──────────────────────────────────────────────────────────────

    history = histories.setdefault(user_id, [])
    history.append({"role": "user", "content": text})

    if _count_history_tokens(history) > _MAX_HISTORY_TOKENS:
        while len(history) > 4 and _count_history_tokens(history) > _MAX_HISTORY_TOKENS:
            history = history[1:]
        histories[user_id] = history

    stop_typing = asyncio.Event()
    typing_task = asyncio.create_task(_keep_typing(message.chat.id, stop_typing))
    try:
        response = await run_llm(history, user_id, message.chat.id)
        response = _clean_response(response)
    except openai.APITimeoutError:
        response = "Сервер думает слишком долго — попробуй повторить запрос."
    except Exception as e:
        response = f"Ошибка: {e}"
    finally:
        stop_typing.set()
        typing_task.cancel()
        try:
            await typing_task
        except asyncio.CancelledError:
            pass

    # Ядерная страховка: если DSML всё ещё есть — не отправлять мусор
    if _has_dsml(response):
        response = _strip_dsml(response)
    if _has_dsml(response):
        response = "Ищу информацию... попробуй повторить запрос."

    try:
        from rule_engine import apply_rules
        response = await apply_rules(response, text, "assistant", ai_client)
    except Exception:
        pass

    history.append({"role": "assistant", "content": response})

    if response and response != "(пустой ответ)":
        for i in range(0, len(response), 4000):
            await message.answer(response[i:i+4000])

# ── Форматирование расписания ─────────────────────────────────
def _format_schedule(date_iso: str, cal_events: list, reminders: list, header: str = "") -> str:
    date_label = datetime.strptime(date_iso, "%Y-%m-%d").strftime("%d.%m.%Y")
    lines = [header or f"📆 *План на {date_label}:*\n"]
    if cal_events:
        lines.append("📅 *Календарь:*")
        for ev in cal_events:
            start = ev["start"].get("dateTime", ev["start"].get("date", ""))
            prefix = datetime.fromisoformat(start).strftime("%H:%M") if "T" in start else "весь день"
            lines.append(f"• {prefix} — {ev.get('summary', '(без названия)')}")
    else:
        lines.append("📅 *Календарь:* событий нет")
    if reminders:
        lines.append("\n⏰ *Напоминания:*")
        for r in reminders:
            lines.append(f"• {r['remind_at'][11:16]} — {r['text']}")
    if not cal_events and not reminders:
        lines.append("Ничего не запланировано 🎉")
    return "\n".join(lines)


# ── Утренний дайджест ─────────────────────────────────────────
async def morning_briefing():
    """Отправляет план на день в 09:00 МСК каждому пользователю."""
    for user_id in config.ALLOWED_USER_IDS:
        try:
            today_iso = datetime.now(_MSK).strftime("%Y-%m-%d")
            today_label = datetime.now(_MSK).strftime("%d.%m.%Y")
            try:
                cal_events = calendar.get_schedule_for_date(config.GOOGLE_CALENDAR_ID, config.SERVICE_ACCOUNT_JSON, today_iso)
            except Exception as e:
                cal_events = []
            reminders = rem_tool.get_reminders_for_date(user_id, today_iso)
            text = _format_schedule(today_iso, cal_events, reminders, header=f"☀️ *Доброе утро! План на {today_label}:*\n")
            await bot.send_message(user_id, text, parse_mode="Markdown")
        except Exception as e:
            print(f"[morning_briefing] ошибка для {user_id}: {e}")


# ── Еженедельный обзор (понедельник 09:00) ────────────────────
async def weekly_review():
    """Отправляет обзор прошедшей недели каждый понедельник в 09:00 МСК."""
    for user_id in config.ALLOWED_USER_IDS:
        try:
            from tools.db import get_conn as _db_conn
            conn = _db_conn()

            # Заметки за последние 7 дней
            rows_notes = conn.execute(
                "SELECT id, text, created_at FROM notes WHERE created_at >= datetime('now', '-7 days') ORDER BY id DESC"
            ).fetchall()

            # Идеи в очереди (категория idea/ideas)
            rows_ideas = conn.execute(
                "SELECT id, text, created_at FROM memory WHERE category IN ('idea','ideas') ORDER BY id DESC LIMIT 20"
            ).fetchall()

            # Выполненные напоминания за неделю
            rows_rem = conn.execute(
                """SELECT text, remind_at FROM reminders
                   WHERE user_id=? AND done=1
                   AND remind_at >= datetime('now', '-7 days')
                   ORDER BY remind_at DESC""",
                (user_id,)
            ).fetchall()
            conn.close()

            lines = ["📊 *Обзор недели*\n"]

            if rows_notes:
                lines.append(f"📝 *Заметки за неделю ({len(rows_notes)} шт):*")
                for r in rows_notes:
                    dt = r["created_at"][:10]
                    lines.append(f"  • ({dt}) {r['text'][:120]}")
                lines.append("")

            if rows_ideas:
                lines.append(f"💡 *Идеи в очереди ({len(rows_ideas)} шт):*")
                for r in rows_ideas:
                    lines.append(f"  • {r['text'][:120]}")
                lines.append("")

            if rows_rem:
                lines.append(f"✅ *Выполнено за неделю ({len(rows_rem)} задач):*")
                for r in rows_rem[:10]:
                    lines.append(f"  • {r['text']}")
                lines.append("")

            if len(lines) == 1:
                lines.append("За прошедшую неделю нет заметок, идей и выполненных задач.")

            await bot.send_message(user_id, "\n".join(lines), parse_mode="Markdown")
        except Exception as e:
            print(f"[weekly_review] ошибка для {user_id}: {e}")




# ── Вечерний дайджест договорённостей ────────────────────────
async def _send_agreements_digest(user_id: int, items: list, header: str):
    """Отправляет список договорённостей отдельными сообщениями с кнопками."""
    from tools.db import mark_digest_sent
    await bot.send_message(user_id, header)
    sent_ids = []
    for agr in items:
        kb = _agreement_kb(agr["id"])
        await bot.send_message(user_id, f"📌 {agr['text']}", reply_markup=kb)
        sent_ids.append(agr["id"])
    if sent_ids:
        mark_digest_sent(sent_ids)


async def send_evening_digest():
    """22:00 МСК — отправляет договорённости за сегодня."""
    from tools.db import get_pending_agreements
    for user_id in config.ALLOWED_USER_IDS:
        try:
            items = get_pending_agreements(only_today=True)
            if not items:
                continue
            header = f"📋 *Итоги дня — {len(items)} договорённост{'ь' if len(items)==1 else 'и' if len(items)<5 else 'ей'}:*"
            await _send_agreements_digest(user_id, items, header)
        except Exception as e:
            logging.error(f"[EveningDigest] ошибка для {user_id}: {e}")


async def send_morning_carryover():
    """10:00 МСК — повторяет вчерашние неразобранные договорённости."""
    from tools.db import get_pending_agreements
    for user_id in config.ALLOWED_USER_IDS:
        try:
            items = get_pending_agreements(only_today=False)
            if not items:
                continue
            header = f"⏰ *Не разобрано вчера — {len(items)} пункт{'а' if len(items)<5 else 'ов'}:*"
            await _send_agreements_digest(user_id, items, header)
        except Exception as e:
            logging.error(f"[MorningCarryover] ошибка для {user_id}: {e}")


# ── Команда /schedule ─────────────────────────────────────────
@dp.message(Command("schedule"))
async def cmd_schedule(message: Message):
    if not is_allowed(message.from_user.id):
        return
    today_iso = datetime.now(_MSK).strftime("%Y-%m-%d")
    try:
        cal_events = calendar.get_schedule_for_date(config.GOOGLE_CALENDAR_ID, config.SERVICE_ACCOUNT_JSON, today_iso)
    except Exception:
        cal_events = []
    reminders = rem_tool.get_reminders_for_date(message.from_user.id, today_iso)
    await message.answer(_format_schedule(today_iso, cal_events, reminders), parse_mode="Markdown")


# ── Команда /digest — ручной запуск дайджеста ─────────────────
@dp.message(Command("digest"))
async def cmd_digest(message: Message):
    if not is_allowed(message.from_user.id):
        return
    from tools.db import get_pending_agreements
    items_today = get_pending_agreements(only_today=True)
    items_old = get_pending_agreements(only_today=False)
    total = items_today + items_old
    if not total:
        await message.answer("📋 Нет невыполненных договорённостей.")
        return
    if items_old:
        header = f"⏰ *Не разобрано ранее — {len(items_old)} пункт{'а' if len(items_old)<5 else 'ов'}:*"
        await _send_agreements_digest(message.from_user.id, items_old, header)
    if items_today:
        header = f"📋 *Сегодня — {len(items_today)} договорённост{'ь' if len(items_today)==1 else 'и' if len(items_today)<5 else 'ей'}:*"
        await _send_agreements_digest(message.from_user.id, items_today, header)


# ── Запуск ────────────────────────────────────────────────────
async def main():
    init_db()
    mem_tool._init_memory_table()
    rem_tool.init_scheduler(bot, scheduler)
    scheduler.add_job(
        morning_briefing, "cron",
        hour=9, minute=0,
        timezone="Europe/Moscow",
        id="morning_briefing",
        replace_existing=True,
    )
    scheduler.add_job(
        send_evening_digest, "cron",
        hour=22, minute=0,
        timezone="Europe/Moscow",
        id="evening_digest",
        replace_existing=True,
    )
    scheduler.add_job(
        send_morning_carryover, "cron",
        hour=10, minute=0,
        timezone="Europe/Moscow",
        id="morning_carryover",
        replace_existing=True,
    )
    scheduler.start()
    # Удаляем webhook при старте — защита от конфликта с внешними сервисами
    try:
        await bot.delete_webhook(drop_pending_updates=True)
    except Exception as e:
        print(f"[webhook] delete error: {e}")
    # Меню команд (подсказки при вводе /)
    from aiogram.types import BotCommand
    await bot.set_my_commands([
        BotCommand(command="help",      description="Шпаргалка — все команды и примеры"),
        BotCommand(command="reminders", description="Активные напоминания"),
        BotCommand(command="notes",     description="Последние заметки"),
        BotCommand(command="memory",    description="Сохранённые факты"),
        BotCommand(command="schedule",  description="Расписание на сегодня"),
        BotCommand(command="digest",    description="Договорённости — показать все"),
        BotCommand(command="reset",     description="Очистить контекст диалога"),
    ])
    print(f"[{datetime.now(_MSK).strftime('%H:%M:%S')} МСК] Бот запущен.")
    await dp.start_polling(bot, allowed_updates=["message", "callback_query"])

if __name__ == "__main__":
    asyncio.run(main())
