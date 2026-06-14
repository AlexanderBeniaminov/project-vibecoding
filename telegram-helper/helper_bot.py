#!/usr/bin/env python3
"""
helper_bot.py — @pomoshniknamac_bot «Помощник»
Работает 24/7 на VPS. Функции: задачи Губаха, CRM, таблицы, email, фото, AI-чат.
"""
import asyncio
import base64
import io
import logging
import os
import re
import sys
import json
import imaplib
import smtplib
import sqlite3
import tempfile
import email as email_lib
import email.header
import email.utils
from datetime import datetime, timedelta
from email.mime.text import MIMEText
from email.header import Header
from pathlib import Path
from zoneinfo import ZoneInfo

import openai
import requests
from bs4 import BeautifulSoup
from ddgs import DDGS
from aiogram import Bot, Dispatcher, F
from aiogram.filters import Command
from aiogram.types import Message, InlineKeyboardMarkup, InlineKeyboardButton, CallbackQuery
from groq import Groq
from apscheduler.schedulers.asyncio import AsyncIOScheduler

import config
from smart_search import COMMERCIAL_RE, detect_type, get_clarification, smart_search_and_answer
from shopping import search_all_stores, format_results as format_shopping

# ── Инициализация ─────────────────────────────────────────────────
_MSK = ZoneInfo("Europe/Moscow")
bot = Bot(token=config.TELEGRAM_TOKEN)
dp = Dispatcher()
scheduler = AsyncIOScheduler(timezone="Europe/Moscow")

ai_client = openai.AsyncOpenAI(
    base_url=config.ROUTERAI_BASE_URL,
    api_key=config.ROUTERAI_API_KEY,
)

_HOTELIER_DB = Path(config.DATA_DIR) / "hotelier.db"

histories: dict[int, list] = {}
_pending_reformat: dict[int, dict] = {}  # {user_id: {instruction, original}}
_last_message_time: dict[int, float] = {}
_CONTEXT_TTL = 300  # 5 минут — после паузы контекст считается новым
_pending_searches: dict[int, str] = {}   # частичный коммерческий запрос ожидает уточнения

# ── Проверка доступа ──────────────────────────────────────────────
def is_allowed(user_id: int) -> bool:
    return user_id in config.ALLOWED_USER_IDS

# ── Системный промпт ──────────────────────────────────────────────
def _system_prompt() -> str:
    now = datetime.now(_MSK).strftime("%d.%m.%Y %H:%M")
    base_prompt = (
        f"Ты бизнес-ассистент Александра Бениаминова. Сейчас {now} МСК. Язык: русский.\n\n"
        "МОДЕЛЬ: Если задача требует глубокого анализа, составления отчёта или сложного "
        "рассуждения — сам предложи: «Это лучше решит Claude Sonnet 4.6 — переключить?»\n\n"
        "СХЕМА РЕШЕНИЙ — выполняй первое подходящее:\n\n"
        "① Имя из команды + задача → add_team_task.\n"
        "   Команда: Виктор/Витя, Евгения/Женя, Надежда/Надя, Управляющий, Тех.директор.\n"
        "   При нечётком имени — сначала уточни: «Имеешь в виду Евгению?»\n"
        "   ОБЯЗАТЕЛЬНО уточни дедлайн если не указан — задачу без срока не ставь.\n"
        "   Если задача неясна — уточни до конца перед записью.\n"
        "   Голосовое сообщение с задачей обрабатывай так же как текст — уточни если неясно.\n"
        "   После записи задачи → спроси: «Поставить напоминание проверить [имя] к [дедлайн]?»\n\n"
        "② Вопрос про выручку / загрузку / гостей / ADR / RevPar / фудкост / Монблан / Губаха → query_sheets.\n"
        "   Ответ: метрика + значение + период — одной строкой.\n"
        "   Нет данных за период → «Данных за [период] нет». Не ищи замену. Не угадывай.\n\n"
        "③ «задачи [имя]» / «что висит у [имя]» / «что у [имя] в работе» →\n"
        "   query_sheets фильтр по исполнителю. Выводи: задача — дедлайн — статус.\n\n"
        "④ «что по [имя]» / /crm [имя] → crm_get.\n"
        "   Триггеры авто-сохранения в CRM из текста:\n"
        "   «договорились с», «встреча с», «[Имя] хочет», «[Имя] должен», «[Имя] обещал».\n"
        "   При сохранении даты следующего контакта → спроси:\n"
        "   «Поставить напоминание связаться с [имя] [дата]?»\n\n"
        "⑤ /mail / «почта» / «письма» / «что пришло» → mail_digest.\n"
        "   Только папка «Входящие». Фильтруй рассылки, автоуведомления, no-reply — не показывай.\n"
        "   Показывай: от кого, тема, требует ли ответа.\n\n"
        "⑥ «напиши письмо [кому] о [теме]» / «отправь письмо» → compose_email.\n"
        "   Если нужны цифры или даты → сначала вызови query_sheets, подтяни данные,\n"
        "   затем составляй письмо с реальными числами.\n"
        "   Показать черновик → ждать правок или ✅ → только тогда отправить.\n"
        "   Контакты: Виктор (karpenko@entens.ru), Катя (info@entens.ru), Алексей (ap@entens.ru).\n"
        "   НЕЛЬЗЯ отправлять без явного подтверждения.\n\n"
        "⑦ Фото → analyze_photo.\n"
        "   Чек → поставщик, сумма, дата, позиции.\n"
        "   Накладная → поставщик, позиции, итоговая сумма.\n"
        "   Договор → стороны, предмет, сумма, срок, ключевые условия.\n"
        "   Список / текст → транскрибируй дословно.\n"
        "   Если тип = «проблема на объекте»:\n"
        "     1. Опиши проблему кратко.\n"
        "     2. Спроси: «Кому поставить задачу? Виктор / Евгения / Надежда / Управляющий / Тех.директор»\n"
        "     3. После ответа → уточни дедлайн → add_team_task (поле фото не заполняй).\n\n"
        "⑧ «запомни что...» / важный факт о человеке, компании, договорённости → remember_fact.\n"
        "   Категории не навязывай — используй тег по контексту. Память общая с Напоминатором.\n\n"
        "⑨ Всё остальное (не бизнес-задача) → ответь из общих знаний коротко.\n"
        "   Если нужна актуальная информация (цены, адреса, расписания, места) —\n"
        "   напиши ровно одно слово: ПОИСК\n\n"
        "ЖЁСТКИЕ ЗАПРЕТЫ:\n"
        "- НЕЛЬЗЯ ставить задачу без дедлайна.\n"
        "- НЕЛЬЗЯ выдумывать данные из таблиц.\n"
        "- НЕЛЬЗЯ записывать задачу с нечётким именем без уточнения.\n"
        "- НЕЛЬЗЯ отправлять письмо без подтверждения.\n"
        "- НЕЛЬЗЯ предлагать альтернативный период если данных нет — говори прямо.\n"
        "- НЕЛЬЗЯ отказываться от ответа словами «вне моих компетенций» — либо ответь, либо напиши ПОИСК.\n\n"
        "СТИЛЬ: коротко и конкретно. Уточняй неполные запросы — не угадывай."
    )
    try:
        sys.path.insert(0, "/home/parser/bots/shared")
        from rule_engine import get_system_addons as _get_addons
        addons = _get_addons("helper")
        if addons:
            base_prompt += f"\n\nПРАВИЛА ПОЛЬЗОВАТЕЛЯ:\n{addons}"
    except Exception:
        pass
    return base_prompt

# ── Нормализация ASR-ошибок Whisper ──────────────────────────────
_ASR_FIXES = [
    # Монблан — частые ошибки распознавания
    (r'\bмаун\w*лан\w*', 'монблан'),
    (r'\bман\w*лан\w*', 'монблан'),
    (r'\bмонб\w{2,}\b', 'монблан'),
    (r'\bmont\s*blanc\b', 'монблан'),
]
def _fix_asr(text: str) -> str:
    result = text
    for pattern, replacement in _ASR_FIXES:
        result = re.sub(pattern, replacement, result, flags=re.IGNORECASE)
    return result

# ── Очистка ответа ────────────────────────────────────────────────
def _clean(text: str) -> str:
    text = re.sub(r'<\|?\s*(?:think|thinking|DSML)\s*\|?>.*?<\|?\s*/?\s*(?:think|thinking|DSML)\s*\|?>', '', text, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text.strip() or "(пустой ответ)"

# ── Транскрипция голоса ───────────────────────────────────────────
async def transcribe_voice(message: Message) -> str:
    voice = message.voice or message.audio
    if not voice:
        return ""
    file = await bot.get_file(voice.file_id)
    file_bytes = await bot.download_file(file.file_path)
    groq_client = Groq(api_key=config.GROQ_API_KEY)
    try:
        tr = groq_client.audio.transcriptions.create(
            file=("voice.ogg", file_bytes.read()),
            model="whisper-large-v3-turbo",
            language="ru",
        )
        return tr.text.strip()
    except Exception as e:
        return f"[Не удалось распознать: {e}]"

# ── Безопасная отправка (fallback без Markdown) ───────────────────
async def _safe_answer(message: Message, text: str, parse_mode: str = "Markdown"):
    try:
        await message.answer(text, parse_mode=parse_mode)
    except Exception:
        # Если Markdown сломан — отправляем plain text
        await message.answer(text, parse_mode=None)


async def _send_smart_result(message: Message, result: tuple) -> None:
    """Отправляет ответ умного поиска (текст, ссылки встроены прямо в него)."""
    text, _links = result
    await _safe_answer(message, text, parse_mode=None)

# ── Typing indicator ──────────────────────────────────────────────
async def _keep_typing(chat_id: int, stop: asyncio.Event):
    while not stop.is_set():
        await bot.send_chat_action(chat_id, "typing")
        try:
            await asyncio.wait_for(stop.wait(), timeout=4)
        except asyncio.TimeoutError:
            pass

# ── LLM вызов ────────────────────────────────────────────────────
async def run_llm(messages: list, model: str = None) -> str:
    m = model or config.MODEL
    resp = await ai_client.chat.completions.create(
        model=m,
        messages=messages,
        max_tokens=2000,
        temperature=0.5,
    )
    msg = resp.choices[0].message
    content = (msg.content or "").strip()
    if not content and hasattr(msg, "reasoning"):
        content = (msg.reasoning or "")[-600:]
    return _clean(content)


# ══════════════════════════════════════════════════════════════════
# БЛОК 1: ЗАДАЧИ КОМАНДЕ ГУБАХА → Google Sheets
# ══════════════════════════════════════════════════════════════════
_TEAM_ALIASES = {
    "виктор": "Виктор", "viktor": "Виктор",
    "евгения": "Евгения", "женя": "Евгения",
    "надежда": "Надежда", "надя": "Надежда",
    "управляющий": "Управляющий",
    "техдиректор": "Тех.директор", "тех.директор": "Тех.директор",
    "технический директор": "Тех.директор",
}

def _normalize_executor(name: str) -> str:
    return _TEAM_ALIASES.get(name.strip().lower(), name.strip().capitalize())

def _add_gubaha_task_sync(executor: str, task: str, deadline: str = "") -> str:
    from google.oauth2.service_account import Credentials
    from googleapiclient.discovery import build
    creds = Credentials.from_service_account_file(
        config.PERSONAL_SA, scopes=["https://www.googleapis.com/auth/spreadsheets"]
    )
    svc = build("sheets", "v4", credentials=creds, cache_discovery=False)
    executor_clean = _normalize_executor(executor)
    row = [executor_clean, "Устно", task.strip(), "", "", deadline]
    svc.spreadsheets().values().append(
        spreadsheetId=config.STRATEGY_SHEET_ID,
        range="'Задачи недели'!A:F",
        valueInputOption="USER_ENTERED",
        insertDataOption="INSERT_ROWS",
        body={"values": [row]},
    ).execute()
    deadline_str = f" (срок: {deadline})" if deadline else ""
    return f"✅ Задача добавлена: *{executor_clean}* — {task}{deadline_str}"

async def add_gubaha_task(executor: str, task: str, deadline: str = "") -> str:
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, _add_gubaha_task_sync, executor, task, deadline)

# Детектор задачи в тексте
_TASK_RE = re.compile(
    r"(виктор[уа]?|евгени[ия]|жен[ея]|надежд[еы]?|над[еяь]|управляющему?|техдиректор[уа]?|тех\.директор[уа]?)"
    r"[\s,]+задач[ауие]?:?\s*(.+?)(?:\s+(?:до|срок|дедлайн)\s+(.+))?$",
    re.IGNORECASE | re.DOTALL
)


# ══════════════════════════════════════════════════════════════════
# БЛОК 2: CRM
# ══════════════════════════════════════════════════════════════════
_CRM_DB = Path(config.DATA_DIR) / "crm.db"

def _crm_conn():
    _CRM_DB.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(_CRM_DB))
    conn.row_factory = sqlite3.Row
    conn.execute("""CREATE TABLE IF NOT EXISTS contacts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        note TEXT NOT NULL,
        created_at TEXT NOT NULL DEFAULT (datetime('now', 'localtime'))
    )""")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_name ON contacts(name COLLATE NOCASE)")
    conn.commit()
    return conn

def crm_save(name: str, note: str) -> str:
    name_clean = name.strip().title()
    conn = _crm_conn()
    conn.execute("INSERT INTO contacts (name, note) VALUES (?, ?)", (name_clean, note.strip()))
    conn.commit()
    conn.close()
    return f"✅ CRM: сохранено для *{name_clean}*\n_{note.strip()}_"

def crm_get(name: str) -> str:
    conn = _crm_conn()
    rows = conn.execute(
        "SELECT note, created_at FROM contacts WHERE name LIKE ? ORDER BY id DESC LIMIT 8",
        (f"%{name.strip().title()}%",)
    ).fetchall()
    conn.close()
    if not rows:
        return f"По контакту «{name}» ничего не найдено."
    lines = [f"📇 *{name.strip().title()}*:"]
    for r in rows:
        lines.append(f"  • ({r['created_at'][:16]}) {r['note']}")
    return "\n".join(lines)

# Авто-детектор CRM фраз
_CRM_TRIGGERS = ["договорились с ", "сказал ", " хочет ", " должен ", "позвонить ", "встреча с "]


# ══════════════════════════════════════════════════════════════════
# БЛОК 3: ДАННЫЕ ИЗ ТАБЛИЦ
# ══════════════════════════════════════════════════════════════════

# ID таблиц (из Drive-папки 1mfJEc9_XefwJQUWbfnMLOFf7EKSKEJgL)
_MONBLAN_ID    = "1Wcvn2mJFgOfcdm3mUQpYLoU92H3_bhGUJA_NnBwbDNI"  # Отчет из IIKO
_GOG_MONTHLY_ID = "1ADBN7fecyqFvEG6igJfVfwG2vso6oW742y3-ai9ps8c" # ЕжеМесячный отчет ГОГ
_GOG_WEEKLY_ID  = "1Ohm7tst750zDzSeIewJFj_cPC6vl0-5J0UiuNfZvY_k" # ЕжеНедельный отчет ГОГ

_MONTH_NUM = {
    "январ": 1, "феврал": 2, "март": 3, "марта": 3, "апрел": 4,
    "май": 5, "мая": 5, "июн": 6, "июл": 7,
    "август": 8, "сентябр": 9, "октябр": 10, "ноябр": 11, "декабр": 12,
}
_MONTH_RU = {
    1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
    5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
    9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь",
}

def _idx_to_col(i: int) -> str:
    """Индекс колонки → буква (0=A, 26=AA, ...)"""
    if i < 26:
        return chr(65 + i)
    return chr(64 + i // 26) + chr(65 + i % 26)

def _parse_month_year(q: str) -> tuple:
    """Извлекает (месяц: int, год: int) из текста запроса."""
    year = datetime.now().year
    # Год: "26 года", "2026", "25", "2025"
    y_m = re.search(r'\b(20(\d{2})|(\d{2}))\s*(?:год|г\.?)?', q)
    if y_m:
        raw = y_m.group(1)
        y = int(raw)
        year = 2000 + y if y < 100 else y
    # Месяц
    for prefix, num in _MONTH_NUM.items():
        if prefix in q:
            return (num, year)
    return (None, year)

def _sheets_svc(sa_path: str):
    from google.oauth2.service_account import Credentials
    from googleapiclient.discovery import build
    creds = Credentials.from_service_account_file(
        sa_path, scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"]
    )
    return build("sheets", "v4", credentials=creds, cache_discovery=False)

def _read_range(sa_path: str, sheet_id: str, range_str: str) -> list:
    return _sheets_svc(sa_path).spreadsheets().values().get(
        spreadsheetId=sheet_id, range=range_str
    ).execute().get("values", [])

def _direct_lookup(rows_kv: list, question: str, context: str) -> str | None:
    """Прямой поиск без LLM — для типовых запросов по выручке/загрузке/гостям."""
    q = question.lower()
    result_lines = []

    # Выручка — ищем строку "Выручка всего"
    if any(k in q for k in ["выручк", "доход"]):
        for row in rows_kv:
            if re.match(r"выручка всего", row, re.IGNORECASE) or re.match(r"доход", row, re.IGNORECASE):
                if ":" in row:
                    label, value = row.split(":", 1)
                    return f"{label.strip()} {context}: {value.strip()} руб."

    # Загрузка
    if "загрузк" in q:
        for row in rows_kv:
            if re.match(r"загрузка", row, re.IGNORECASE) and ":" in row:
                label, value = row.split(":", 1)
                return f"{label.strip()} {context}: {value.strip()}"

    # Гости / количество гостей
    if any(k in q for k in ["гост", "посетит", "человек"]):
        for row in rows_kv:
            if re.match(r"(кол-во гостей|гост|посетит)", row, re.IGNORECASE) and ":" in row:
                label, value = row.split(":", 1)
                return f"{label.strip()} {context}: {value.strip()}"

    # ADR / RevPar
    if "adr" in q:
        for row in rows_kv:
            if re.match(r"adr", row, re.IGNORECASE) and ":" in row:
                label, value = row.split(":", 1)
                return f"{label.strip()} {context}: {value.strip()} руб."
    if "revpar" in q:
        for row in rows_kv:
            if re.match(r"revpar", row, re.IGNORECASE) and ":" in row:
                label, value = row.split(":", 1)
                return f"{label.strip()} {context}: {value.strip()} руб."

    return None   # не нашли — идём в LLM

def _extract_answer(text: str, question: str = "") -> str:
    """Вырезает финальный ответ из текста с рассуждениями DeepSeek."""
    if not text:
        return ""
    # Убираем кавычки и артефакты в начале
    text = text.strip().strip('"«»').strip()
    # Если ответ — эхо вопроса (похоже на повтор), возвращаем пусто
    if question:
        q_words = set(re.findall(r'\w{4,}', question.lower()))
        a_words = set(re.findall(r'\w{4,}', text.lower()))
        overlap = len(q_words & a_words) / max(len(q_words), 1)
        if overlap > 0.6 and not re.search(r'\d{3,}', text):
            return "Нет данных за этот период."
    # Ищем строку с числом (ответ с цифрой) или вопросом
    for line in text.split("\n"):
        line = line.strip().strip('"«»').strip()
        if not line:
            continue
        if re.search(r'\d{3,}', line):   # есть число ≥ 3 цифр — это цифра выручки
            return line
        if line.endswith("?"):            # уточняющий вопрос
            return line
    # Fallback — первая непустая строка
    for line in text.split("\n"):
        line = line.strip()
        if line:
            return line[:200]
    return text[:200]

# Загрузка данных (sync) — возвращает (rows_kv, context)
def _monblan_monthly_load(month: int, year: int) -> tuple:
    svc = _sheets_svc(config.MONBLAN_SA)
    header = svc.spreadsheets().values().get(
        spreadsheetId=_MONBLAN_ID, range="ЕжеМесячный!A2:AZ2"
    ).execute().get("values", [[]])[0]
    target = f"{year}-{month:02d}"
    if target not in header:
        return (None, f"Данных Монблан за {_MONTH_RU.get(month, month)} {year} нет в таблице.")
    col_idx = header.index(target)
    col_letter = _idx_to_col(col_idx)
    rows = svc.spreadsheets().values().get(
        spreadsheetId=_MONBLAN_ID, range=f"ЕжеМесячный!A3:{col_letter}60"
    ).execute().get("values", [])
    kv = []
    for row in rows:
        label = row[0].strip() if row and row[0] else ""
        value = row[col_idx].strip() if len(row) > col_idx and row[col_idx] else ""
        if label and value and not label.startswith("%"):
            kv.append(f"{label}: {value}")
    return (kv, f"Монблан {_MONTH_RU.get(month, month)} {year}")

def _gog_monthly_load(month: int, year: int) -> tuple:
    short_year = str(year)[-2:]
    sheet_name = f"{_MONTH_RU.get(month, month)} {short_year}"
    try:
        rows = _read_range(config.AIHOTEL_SA, _GOG_MONTHLY_ID, f"'{sheet_name}'!A1:F30")
    except Exception:
        return (None, f"Данных ГОГ за {sheet_name} нет в таблице.")
    kv = []
    for row in rows:
        if len(row) >= 2 and row[1]:
            parts = [row[1]]
            if len(row) > 2 and row[2]: parts.append(f"{year-1}: {row[2]}")
            if len(row) > 4 and row[4]: parts.append(f"{year}: {row[4]}")
            kv.append(" | ".join(parts))
    return (kv, f"ГОГ Губаха {sheet_name}")

def _load_sheets_data(question: str) -> tuple:
    """Sync: роутинг и загрузка данных. Возвращает (rows_kv | None, context_or_error)."""
    q = question.lower()
    month, year = _parse_month_year(q)
    is_monblan = any(k in q for k in ["монблан", "monblan", "кафе", "ресторан"])
    is_gog     = any(k in q for k in ["губаха", "отель", "хостел", "коттедж", "гог", "номер", "загрузк", "заезд", "бронь"])
    is_weekly  = any(k in q for k in ["недел", "неделю", "неделя"])
    is_daily   = any(k in q for k in ["сегодня", "вчера", "день", "дневн"])
    is_tasks   = any(k in q for k in ["задач", "kpi", "стратег"])

    if is_tasks:
        rows = _read_range(config.PERSONAL_SA, config.STRATEGY_SHEET_ID, "'Задачи недели'!A1:H50")
        kv = [" | ".join(str(c) for c in row if c) for row in rows if any(row)]
        return (kv, "Задачи недели Губаха")

    if is_monblan:
        if month and not is_weekly and not is_daily:
            return _monblan_monthly_load(month, year)
        elif is_daily:
            rows = _read_range(config.MONBLAN_SA, _MONBLAN_ID, "ЕжеДневно!A1:AZ15")
            kv = [" | ".join(str(c) for c in row if c) for row in rows if any(row)]
            return (kv, "Монблан ежедневно")
        else:
            rows = _read_range(config.MONBLAN_SA, _MONBLAN_ID, "ЕжеНедельно!A1:CZ30")
            kv = [" | ".join(str(c) for c in row[:15] if c) for row in rows if any(row)]
            return (kv, "Монблан еженедельно")

    if is_gog or (not is_monblan and month):
        if month and not is_weekly:
            return _gog_monthly_load(month, year)
        else:
            rows = _read_range(config.AIHOTEL_SA, _GOG_WEEKLY_ID, "Дайджест!A1:D50")
            kv = [" | ".join(str(c) for c in row if c) for row in rows if any(row)]
            return (kv, "ГОГ Губаха еженедельный дайджест")

    return (None, "Уточни: это про Монблан (кафе) или про Губаха (отель/коттеджи/хостел)?")

async def query_sheets(question: str) -> str:
    """Async: грузит данные → прямой поиск → если не нашли, спрашивает Claude."""
    loop = asyncio.get_event_loop()
    rows_kv, context = await loop.run_in_executor(None, _load_sheets_data, question)

    if rows_kv is None:
        return context

    # 1. Сначала прямой поиск без LLM (быстро и точно)
    direct = _direct_lookup(rows_kv, question, context)
    if direct:
        return direct

    # 2. Fallback — Claude для сложных/составных вопросов
    today = datetime.now(_MSK).strftime("%d.%m.%Y")
    table_str = "\n".join(rows_kv[:60])
    resp = await ai_client.chat.completions.create(
        model=config.VISION_MODEL,
        messages=[
            {"role": "system", "content": (
                f"Данные из таблицы за {context} (уже загружены, период точный).\n"
                "Ответь на вопрос одной строкой: метрика и число. Пример: Кухня март 2026: 2 564 170 руб.\n"
                "Только цифра, никаких объяснений."
            )},
            {"role": "user", "content": f"{table_str}\n\n{question}\n\nОтвет:"}
        ],
        max_tokens=60,
        temperature=0.0,
    )
    raw = (resp.choices[0].message.content or "").strip()
    return _extract_answer(raw, question) or f"Данные за {context} загружены, но ответ не найден."

_SHEETS_KEYWORDS = ["выручка", "загрузка", "гост", "бронь", "заезд", "фудкост",
                    "монблан", "губаха", "отель", "хостел", "коттедж", "номер", "adr", "revpar"]


# ══════════════════════════════════════════════════════════════════
# БЛОК 4: EMAIL (ab@entens.ru, Mail.ru)
# ══════════════════════════════════════════════════════════════════
_MAIL_CONTACTS = {
    "виктор": "karpenko@entens.ru",
    "алексей": "ap@entens.ru",
    "катя": "info@entens.ru",
    "алексей личная": "aprosandeev@mail.ru",
    "себе": "ab@entens.ru",
}

def _mail_cfg() -> dict:
    return json.loads(Path(config.MAILRU_CFG).read_text())

def _resolve_email(name: str) -> str:
    if "@" in name:
        return name
    key = name.lower().strip()
    for k, v in _MAIL_CONTACTS.items():
        if key in k or k in key:
            return v
    raise ValueError(f"Контакт не найден: {name}")

def _smtp_send(to: str, subject: str, body: str) -> None:
    cfg = _mail_cfg()
    msg = MIMEText(body, "plain", "utf-8")
    msg["From"] = f"{Header(cfg.get('display_name',''), 'utf-8').encode()} <{cfg['email']}>"
    msg["To"] = to
    msg["Subject"] = Header(subject, "utf-8").encode()
    msg["Date"] = email.utils.formatdate(localtime=True)
    with smtplib.SMTP_SSL(cfg["smtp_host"], cfg["smtp_port"]) as server:
        server.login(cfg["email"], cfg["password"])
        server.sendmail(cfg["email"], [to], msg.as_string().encode("utf-8"))

def _imap_connect():
    cfg = _mail_cfg()
    mail = imaplib.IMAP4_SSL(cfg["imap_host"], cfg["imap_port"])
    mail.login(cfg["email"], cfg["password"])
    return mail

def _decode_hdr(raw: str) -> str:
    parts = email.header.decode_header(raw or "")
    return " ".join(
        p.decode(c or "utf-8", errors="replace") if isinstance(p, bytes) else p
        for p, c in parts
    )

def mail_digest_sync() -> str:
    import urllib.request
    mail = _imap_connect()
    mail.select("INBOX")
    since = (datetime.now() - timedelta(days=1)).strftime("%d-%b-%Y")
    _, uids = mail.search(None, f'(SINCE "{since}" UNSEEN)')
    mail_uids = uids[0].split() if uids[0] else []
    if not mail_uids:
        mail.logout()
        return "📭 Нет новых писем за 24 часа."
    summaries = []
    for uid in mail_uids[-15:]:
        _, data = mail.fetch(uid, "(RFC822.HEADER)")
        msg = email_lib.message_from_bytes(data[0][1])
        summaries.append(f"UID:{uid.decode()} От:{_decode_hdr(msg.get('From',''))[:50]} Тема:{_decode_hdr(msg.get('Subject',''))[:70]}")
    mail.logout()
    inbox_text = "\n".join(summaries)
    payload = {
        "model": config.MODEL,
        "messages": [
            {"role": "system", "content": "Помощник. Кратко по-русски."},
            {"role": "user", "content": f"Входящие:\n{inbox_text}\n\nВыдели требующие ответа: от кого, тема, что сделать, UID."}
        ],
        "max_tokens": 1500, "temperature": 0.2,
    }
    data = json.dumps(payload).encode()
    req = urllib.request.Request(f"{config.ROUTERAI_BASE_URL}/chat/completions", data=data,
        headers={"Authorization": f"Bearer {config.ROUTERAI_API_KEY}", "Content-Type": "application/json"}, method="POST")
    with urllib.request.urlopen(req, timeout=60) as resp:
        result = json.loads(resp.read().decode())
    msg_r = result["choices"][0]["message"]
    return f"📬 Дайджест ({len(mail_uids)} непрочитанных):\n\n" + ((msg_r.get("content") or "").strip() or (msg_r.get("reasoning") or "")[-600:])

# Pending email drafts
_pending_drafts: dict[int, dict] = {}

async def compose_email(user_id: int, recipient: str, subject_hint: str, voice_text: str) -> str:
    to_email = _resolve_email(recipient)
    today = datetime.now(_MSK).strftime("%d.%m.%Y")
    prompt = (
        f"Напиши деловое письмо.\nПолучатель: {recipient}\nТема: {subject_hint}\n"
        f"Содержание: {voice_text}\nОтвет — JSON: {{\"subject\": \"...\", \"body\": \"...\"}}"
    )
    resp = await ai_client.chat.completions.create(
        model=config.MODEL,
        messages=[
            {"role": "system", "content": f"Деловая переписка. Отправитель — Александр Бениаминов (ab@entens.ru). Сегодня {today}."},
            {"role": "user", "content": prompt}
        ],
        max_tokens=1500, temperature=0.3,
    )
    msg = resp.choices[0].message
    content = (msg.content or "").strip()
    try:
        parsed = json.loads(content.replace("```json","").replace("```","").strip())
        subject = parsed.get("subject", subject_hint)
        body = parsed.get("body", voice_text)
    except Exception:
        subject, body = subject_hint, content

    _pending_drafts[user_id] = {"to": to_email, "subject": subject, "body": body}
    kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="✅ Отправить", callback_data="mail_send"),
        InlineKeyboardButton(text="❌ Отменить", callback_data="mail_cancel"),
    ]])
    return (f"📧 Черновик:\nКому: {to_email}\nТема: {subject}\n\n{body}", kb)


# ══════════════════════════════════════════════════════════════════
# БЛОК 5: ВЕБ-ПОИСК (DuckDuckGo)
# ══════════════════════════════════════════════════════════════════
_SEARCH_RE = re.compile(
    r"^(найди|поищи|найдите|поищите|погугли|ищи|покажи|расскажи про|что такое|кто такой|узнай)\b",
    re.IGNORECASE
)

def _web_search(query: str, max_results: int = 6) -> str:
    try:
        with DDGS() as ddgs:
            results = list(ddgs.text(query, max_results=max_results))
        if not results:
            return "По этому запросу ничего не найдено."
        lines = []
        for r in results:
            lines.append(f"{r['title']}\n{r['body']}\n{r['href']}")
        return "\n\n".join(lines)
    except Exception as e:
        return f"Ошибка поиска: {e}"

async def web_search_and_answer(query: str) -> str:
    """Ищет в интернете и суммаризирует ответ через LLM."""
    loop = asyncio.get_event_loop()
    raw = await loop.run_in_executor(None, _web_search, query)
    if raw.startswith("Ошибка") or raw.startswith("По этому"):
        return raw
    now = datetime.now(_MSK).strftime("%d.%m.%Y %H:%M")
    messages = [
        {"role": "system", "content": f"Ты помощник. Сейчас {now} МСК. Отвечай коротко по-русски — только факты из результатов поиска."},
        {"role": "user", "content": f"Запрос: {query}\n\nРезультаты поиска:\n{raw}\n\nДай краткий ответ."}
    ]
    return await run_llm(messages)


# ══════════════════════════════════════════════════════════════════
# БЛОК 6: ФОТО (Vision через Claude)
# ══════════════════════════════════════════════════════════════════
async def analyze_photo(message: Message, voice_hint: str = "") -> str:
    photo = message.photo[-1] if message.photo else None
    if not photo:
        return "Фото не найдено."
    file = await bot.get_file(photo.file_id)
    file_bytes = await bot.download_file(file.file_path)
    img_b64 = base64.b64encode(file_bytes.read()).decode("utf-8")

    hint_part = f"\n\nКомментарий пользователя: {voice_hint}" if voice_hint else ""
    user_content = [
        {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{img_b64}"}},
        {"type": "text", "text": (
            "Проанализируй фото. Определи тип: проблема на объекте / документ / чек / накладная / список / другое.\n"
            "Для проблемы: опиши коротко что не так.\n"
            "Для документа/чека: извлеки поставщик, сумму, дату, ключевые условия.\n"
            "Для списка/текста: транскрибируй дословно.\n"
            "Отвечай по-русски, структурированно." + hint_part
        )}
    ]
    resp = await ai_client.chat.completions.create(
        model=config.VISION_MODEL,
        messages=[{"role": "user", "content": user_content}],
        max_tokens=1000,
    )
    return (resp.choices[0].message.content or "").strip()


# ══════════════════════════════════════════════════════════════════
# БЛОК 6: ПРОЕКТНЫЕ ЗАМЕТКИ → Google Sheets
# ══════════════════════════════════════════════════════════════════
_PROJ_NOTES_SHEET_ID = "1aiCKUs-Le-adHSfAOOC2AHRs1vCWzm-E3lucPTbRRkc"
_PROJ_MAP = {
    "губаха": "Губаха", "губахи": "Губаха", "губахе": "Губаха", "губаху": "Губаха",
    "юникорн": "Юникорн", "unicorn": "Юникорн",
    "прочее": "Без проекта", "офис": "Без проекта",
    "разное": "Без проекта", "без проекта": "Без проекта",
}
_PROJ_RE = re.compile(
    r"(?:для\s+)?(?:проект(?:а|е|у)?\s+)?(губах[аиу]?|юникорн|unicorn|прочее?|офис|разное|без\s+проекта)\W+(.+)",
    re.IGNORECASE | re.DOTALL
)

def _save_proj_note_sync(project_raw: str, text: str) -> str:
    from google.oauth2.service_account import Credentials
    from googleapiclient.discovery import build
    project = _PROJ_MAP.get(project_raw.strip().lower(), project_raw.strip().capitalize())
    now = datetime.now()
    creds = Credentials.from_service_account_file(
        config.PERSONAL_SA, scopes=["https://www.googleapis.com/auth/spreadsheets"]
    )
    svc = build("sheets", "v4", credentials=creds, cache_discovery=False)
    meta = svc.spreadsheets().get(spreadsheetId=_PROJ_NOTES_SHEET_ID).execute()
    if project not in [s["properties"]["title"] for s in meta["sheets"]]:
        svc.spreadsheets().batchUpdate(spreadsheetId=_PROJ_NOTES_SHEET_ID, body={
            "requests": [{"addSheet": {"properties": {"title": project}}}]
        }).execute()
        svc.spreadsheets().values().update(
            spreadsheetId=_PROJ_NOTES_SHEET_ID, range=f"'{project}'!A1:C1",
            valueInputOption="USER_ENTERED", body={"values": [["Дата","Время","Мысль / Идея / План"]]}
        ).execute()
    svc.spreadsheets().values().append(
        spreadsheetId=_PROJ_NOTES_SHEET_ID, range=f"'{project}'!A:C",
        valueInputOption="USER_ENTERED", insertDataOption="INSERT_ROWS",
        body={"values": [[now.strftime("%d.%m.%Y"), now.strftime("%H:%M"), text.strip()]]},
    ).execute()
    return f"✅ Записано в *{project}*:\n_{text.strip()}_"


# ══════════════════════════════════════════════════════════════════
# HANDLERS
# ══════════════════════════════════════════════════════════════════
@dp.message(Command("start"))
async def cmd_start(message: Message):
    if not is_allowed(message.from_user.id):
        return
    await message.answer(
        "👋 *Помощник* — работаю 24/7\n\n"
        "Что умею:\n"
        "• Задачи команде Губаха → Sheets\n"
        "• Данные из таблиц голосом\n"
        "• CRM контактов\n"
        "• Почта ab@entens.ru\n"
        "• Фото → анализ/задача\n"
        "• Проектные заметки\n\n"
        "/help — подробная шпаргалка",
        parse_mode="Markdown"
    )

@dp.message(Command("help"))
async def cmd_help(message: Message):
    if not is_allowed(message.from_user.id):
        return
    await message.answer(
        "*Помощник — шпаргалка*\n\n"
        "✈️ *Умный поиск*\n"
        "• Найди авиабилеты в Пермь 22 июня из Москвы\n"
        "• Найди отели в Баку с 1 по 3 ноября\n"
        "→ Бот уточнит если чего-то не хватает, даст ссылки\n\n"
        "🛒 *Шопинг — сравнение цен*\n"
        "• Купить нож для кухни\n"
        "• Где купить молоко Пармалат дешевле\n"
        "• Сколько стоит кофе Lavazza 1кг\n"
        "• Где дешевле: молоко, хлеб, масло\n"
        "→ Даю две таблицы: оптимальная корзина + всё в одном магазине\n\n"
        "📋 *Задачи команде Губаха*\n"
        "• «Задача Виктору: проверить инженера» → в Sheets\n"
        "• Голосовое тоже работает\n\n"
        "📊 *Данные из таблиц*\n"
        "• Выручка Монблан за май\n"
        "• Загрузка отеля за апрель\n\n"
        "👤 *CRM*: /crm Фамилия\n"
        "📬 *Почта*: /mail — дайджест ab@entens.ru\n"
        "📷 *Фото* — анализ чека, задача из скрина\n"
        "🔄 /new — очистить контекст диалога\n\n"
        "📌 Идеи и мысли голосом → сохраняются в Google Sheets",
        parse_mode="Markdown"
    )

@dp.message(Command("new", "reset"))
async def cmd_reset(message: Message):
    if not is_allowed(message.from_user.id):
        return
    histories.pop(message.from_user.id, None)
    await message.answer("🔄 Контекст очищен.")

@dp.message(Command("crm"))
async def cmd_crm(message: Message):
    if not is_allowed(message.from_user.id):
        return
    args = message.text.split(maxsplit=1)
    if len(args) < 2:
        await message.answer("Использование: /crm Богданов")
        return
    await message.answer(crm_get(args[1]), parse_mode="Markdown")

@dp.message(Command("mail"))
async def cmd_mail(message: Message):
    if not is_allowed(message.from_user.id):
        return
    loop = asyncio.get_event_loop()
    await message.answer("📬 Загружаю...")
    result = await loop.run_in_executor(None, mail_digest_sync)
    await message.answer(result, parse_mode="Markdown")

# ── Callback: отправить / отменить письмо ─────────────────────────
@dp.callback_query(F.data == "mail_send")
async def cb_mail_send(call: CallbackQuery):
    user_id = call.from_user.id
    draft = _pending_drafts.pop(user_id, None)
    if not draft:
        await call.answer("Черновик не найден.")
        return
    try:
        loop = asyncio.get_event_loop()
        await loop.run_in_executor(None, _smtp_send, draft["to"], draft["subject"], draft["body"])
        await call.message.edit_text(f"✅ Письмо отправлено!\nКому: {draft['to']}\nТема: {draft['subject']}")
    except Exception as e:
        await call.message.edit_text(f"❌ Ошибка отправки: {e}")

@dp.callback_query(F.data == "mail_cancel")
async def cb_mail_cancel(call: CallbackQuery):
    _pending_drafts.pop(call.from_user.id, None)
    await call.message.edit_text("🗑 Письмо отменено.")

# ── Фото ──────────────────────────────────────────────────────────
_pending_photo_voice: dict[int, Message] = {}

@dp.message(F.photo)
async def handle_photo(message: Message):
    if not is_allowed(message.from_user.id):
        return
    caption = message.caption or ""
    stop = asyncio.Event()
    typing = asyncio.create_task(_keep_typing(message.chat.id, stop))
    try:
        result = await analyze_photo(message, caption)
    finally:
        stop.set(); typing.cancel()
        try: await typing
        except asyncio.CancelledError: pass

    # Если упомянута задача и исполнитель → предложить записать
    task_m = _TASK_RE.search(caption)
    if task_m:
        executor = task_m.group(1)
        task_text = task_m.group(2) or result[:100]
        loop = asyncio.get_event_loop()
        task_result = await loop.run_in_executor(None, _add_gubaha_task_sync, executor, task_text, "")
        await message.answer(f"{result}\n\n{task_result}", parse_mode="Markdown")
    else:
        await message.answer(result, parse_mode="Markdown")

# ── Документы (PDF, Word, Excel, изображения) ────────────────

_MAX_DOC_SIZE = 20 * 1024 * 1024  # 20 МБ — лимит Telegram Bot API


async def _download_doc(doc) -> bytes | None:
    if doc.file_size and doc.file_size > _MAX_DOC_SIZE:
        return None
    tg_file = await bot.get_file(doc.file_id)
    buf = await bot.download_file(tg_file.file_path)
    return buf.read()


async def _analyze_pdf(raw: bytes, hint: str) -> str:
    import pdfplumber
    text = ""
    with pdfplumber.open(io.BytesIO(raw)) as pdf:
        for page in pdf.pages[:15]:
            text += (page.extract_text() or "") + "\n"
    text = text.strip()

    if len(text) > 200:
        prompt = hint or "Проанализируй документ. Выдели: тип, стороны, ключевые условия, сроки, суммы."
        resp = await ai_client.chat.completions.create(
            model=config.MODEL,
            messages=[
                {"role": "system", "content": "Анализируй юридический/деловой документ. Отвечай структурированно по-русски."},
                {"role": "user", "content": f"{prompt}\n\n---\n{text[:12000]}"},
            ],
            max_tokens=2000,
        )
        return _clean((resp.choices[0].message.content or "").strip())

    # Скан → конвертируем в картинки → vision
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        tmp.write(raw); tmp_path = tmp.name
    try:
        from pdf2image import convert_from_path
        images = convert_from_path(tmp_path, dpi=150, first_page=1, last_page=3)
        if not images:
            return "Не удалось извлечь текст из PDF."
        content: list = []
        for img in images[:2]:
            buf = io.BytesIO()
            img.save(buf, format="JPEG", quality=85)
            content.append({"type": "image_url", "image_url": {
                "url": f"data:image/jpeg;base64,{base64.b64encode(buf.getvalue()).decode()}"
            }})
        content.append({"type": "text", "text": hint or "Проанализируй документ. Выдели: тип, стороны, ключевые условия, сроки, суммы."})
        resp = await ai_client.chat.completions.create(
            model=config.MODEL,
            messages=[{"role": "user", "content": content}],
            max_tokens=2000,
        )
        return _clean((resp.choices[0].message.content or "").strip())
    finally:
        os.unlink(tmp_path)


async def _analyze_docx(raw: bytes, hint: str) -> str:
    import docx
    doc = docx.Document(io.BytesIO(raw))
    text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    if not text:
        return "Документ Word пустой или не содержит текста."
    prompt = hint or "Проанализируй документ. Выдели: тип, стороны, ключевые условия, сроки, суммы."
    resp = await ai_client.chat.completions.create(
        model=config.MODEL,
        messages=[
            {"role": "system", "content": "Анализируй документ. Отвечай структурированно по-русски."},
            {"role": "user", "content": f"{prompt}\n\n---\n{text[:12000]}"},
        ],
        max_tokens=2000,
    )
    return _clean((resp.choices[0].message.content or "").strip())


async def _analyze_xlsx(raw: bytes, hint: str) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    parts = []
    for name in wb.sheetnames[:3]:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(max_row=60, values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(c.strip() for c in cells):
                rows.append(" | ".join(cells))
        if rows:
            parts.append(f"Лист «{name}»:\n" + "\n".join(rows[:60]))
    text = "\n\n".join(parts)
    if not text:
        return "Excel-файл пустой или не содержит данных."
    prompt = hint or "Проанализируй таблицу. Что здесь за данные? Выдели ключевые показатели."
    resp = await ai_client.chat.completions.create(
        model=config.MODEL,
        messages=[
            {"role": "system", "content": "Анализируй таблицу. Отвечай по-русски."},
            {"role": "user", "content": f"{prompt}\n\n---\n{text[:8000]}"},
        ],
        max_tokens=1500,
    )
    return _clean((resp.choices[0].message.content or "").strip())


async def _analyze_image_as_doc(raw: bytes, mime: str, hint: str) -> str:
    content = [
        {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{base64.b64encode(raw).decode()}"}},
        {"type": "text", "text": hint or "Проанализируй изображение. Опиши что видишь."},
    ]
    resp = await ai_client.chat.completions.create(
        model=config.MODEL,
        messages=[{"role": "user", "content": content}],
        max_tokens=1500,
    )
    return _clean((resp.choices[0].message.content or "").strip())


@dp.message(F.document)
async def handle_document(message: Message):
    if not is_allowed(message.from_user.id):
        return
    user_id = message.from_user.id
    doc = message.document
    filename = doc.file_name or "document"
    ext = Path(filename).suffix.lower()
    mime = doc.mime_type or ""
    hint = (message.caption or "").strip()

    # Аудио-документы — пропускаем (нет транскрипции в Помощнике)
    if mime.startswith("audio/"):
        return

    mb = round((doc.file_size or 0) / 1024 / 1024, 1)
    if doc.file_size and doc.file_size > _MAX_DOC_SIZE:
        await message.answer(
            f"⚠️ Файл слишком большой ({mb} МБ). "
            f"Telegram позволяет скачивать файлы до 20 МБ.\n"
            f"Попробуй сжать PDF или разбить на части."
        )
        return

    stop = asyncio.Event()
    typing = asyncio.create_task(_keep_typing(message.chat.id, stop))
    try:
        await message.answer(f"📄 Читаю «{filename}»…")
        raw = await _download_doc(doc)
        if raw is None:
            await message.answer("Не удалось скачать файл.")
            return

        if ext == ".pdf" or "pdf" in mime:
            result = await _analyze_pdf(raw, hint)
        elif ext in (".docx", ".doc") or "word" in mime or "officedocument.word" in mime:
            result = await _analyze_docx(raw, hint)
        elif ext in (".xlsx", ".xls") or "excel" in mime or "spreadsheet" in mime:
            result = await _analyze_xlsx(raw, hint)
        elif ext == ".csv":
            text = raw.decode("utf-8", errors="replace")[:8000]
            prompt = hint or "Проанализируй CSV-данные. Выдели ключевые показатели."
            resp = await ai_client.chat.completions.create(
                model=config.MODEL,
                messages=[{"role": "user", "content": f"{prompt}\n\n{text}"}],
                max_tokens=1500,
            )
            result = _clean((resp.choices[0].message.content or "").strip())
        elif mime.startswith("image/") or ext in (".jpg", ".jpeg", ".png", ".webp"):
            result = await _analyze_image_as_doc(raw, mime or "image/jpeg", hint)
        else:
            await message.answer(
                f"❓ Формат «{ext or mime}» не поддерживается.\n"
                f"Поддерживаю: PDF, Word (.docx), Excel (.xlsx), CSV, JPEG/PNG."
            )
            return
    except Exception as e:
        logging.error(f"[doc] ошибка анализа {filename}: {e}")
        result = f"Ошибка при анализе файла: {e}"
    finally:
        stop.set(); typing.cancel()
        try: await typing
        except asyncio.CancelledError: pass

    # Добавляем в историю — можно задавать уточняющие вопросы без повторной отправки
    history = histories.setdefault(user_id, [])
    if not history:
        history.append({"role": "system", "content": _system_prompt()})
    history.append({"role": "user", "content": f"[Файл: {filename}]{chr(10) + hint if hint else ''}"})
    history.append({"role": "assistant", "content": result})
    _last_message_time[user_id] = datetime.now().timestamp()

    for chunk in range(0, len(result), 4000):
        await message.answer(result[chunk:chunk + 4000])


# ── Управление правилами прямо в чате ────────────────────────

def _shared_db():
    if "/home/parser/bots/shared" not in sys.path:
        sys.path.insert(0, "/home/parser/bots/shared")
    import rules_db
    return rules_db

def _shared_engine():
    if "/home/parser/bots/shared" not in sys.path:
        sys.path.insert(0, "/home/parser/bots/shared")
    import rule_engine
    return rule_engine

_BOT_NAME = "helper"

RULE_TYPE_NAMES = {
    "system_addon": "📌 В промпт",
    "reformat": "✏️ Переформат",
    "append": "➕ Дописать в конец",
    "prepend": "⬆️ Вставить в начало",
}


async def _do_reformat(message: Message, text: str) -> None:
    user_id = message.from_user.id
    instruction = re.sub(r'^переделай\s*[—–-]?\s*', '', text, flags=re.IGNORECASE).strip()
    if not instruction:
        await message.answer("Укажи инструкцию: «переделай — пиши короче»")
        return
    original = (message.reply_to_message.text or "").strip()
    if not original:
        await message.answer("Не удалось получить текст исходного сообщения.")
        return
    await message.answer("✏️ Переформатирую...")
    try:
        resp = await ai_client.chat.completions.create(
            model=config.MODEL,
            messages=[
                {"role": "system", "content": "Переформатируй текст согласно инструкции. Верни только результат."},
                {"role": "user", "content": f"Инструкция: {instruction}\n\nТекст:\n{original}"},
            ],
            max_tokens=2000,
        )
        reformatted = (resp.choices[0].message.content or original).strip()
    except Exception as e:
        await message.answer(f"Ошибка: {e}")
        return
    _pending_reformat[user_id] = {"instruction": instruction, "original": original}
    kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="💾 Сохранить как правило", callback_data=f"save_rule:{_BOT_NAME}"),
        InlineKeyboardButton(text="✅ Готово", callback_data="reformat_done"),
    ]])
    await message.answer(f"*Результат:*\n\n{reformatted}\n\n_Сохранить это как правило для следующих ответов?_",
                         parse_mode="Markdown", reply_markup=kb)


@dp.callback_query(F.data.startswith("save_rule:"))
async def cb_save_rule(callback: CallbackQuery):
    bot_name = callback.data.split(":")[1]
    pending = _pending_reformat.pop(callback.from_user.id, None)
    if not pending:
        await callback.answer("Данные истекли.")
        return
    try:
        db = _shared_db()
        rule_id = db.create_rule(bot_name, "reformat", pending["instruction"],
                                 description=f"Авто: {pending['instruction'][:60]}")
        _shared_engine().invalidate_cache(bot_name)
        await callback.message.edit_text(
            f"💾 Правило #{rule_id} сохранено.\nБуду так форматировать следующие ответы."
        )
    except Exception as e:
        await callback.message.edit_text(f"Ошибка сохранения: {e}")
    await callback.answer()


@dp.callback_query(F.data == "reformat_done")
async def cb_reformat_done(callback: CallbackQuery):
    _pending_reformat.pop(callback.from_user.id, None)
    await callback.message.edit_reply_markup(reply_markup=None)
    await callback.answer("Готово.")


@dp.callback_query(F.data.startswith("del_rule:"))
async def cb_del_rule_confirm(callback: CallbackQuery):
    rule_id = int(callback.data.split(":")[1])
    try:
        db = _shared_db()
        ok = db.delete_rule(rule_id)
        _shared_engine().invalidate_cache()
        text = f"🗑 Правило #{rule_id} удалено." if ok else f"Правило #{rule_id} не найдено."
    except Exception as e:
        text = f"Ошибка: {e}"
    await callback.message.edit_text(text)
    await callback.answer()


@dp.callback_query(F.data == "del_cancel")
async def cb_del_cancel(callback: CallbackQuery):
    await callback.message.edit_text("Отменено.")
    await callback.answer()


@dp.message(Command("rules"))
async def cmd_rules(message: Message):
    if not is_allowed(message.from_user.id):
        return
    try:
        rules = _shared_db().list_rules(_BOT_NAME)
    except Exception as e:
        await message.answer(f"Ошибка: {e}")
        return
    if not rules:
        await message.answer("Правил нет. Процитируй мой ответ и напиши «переделай — инструкция».")
        return
    lines = []
    for r in rules:
        status = "✅" if r["active"] else "❌"
        type_label = RULE_TYPE_NAMES.get(r["rule_type"], r["rule_type"])
        desc = r["description"] or r["instruction"][:60]
        lines.append(f"{status} #{r['id']} {type_label}\n    {desc}")
    await message.answer("📋 *Мои правила:*\n\n" + "\n\n".join(lines), parse_mode="Markdown")


@dp.message(Command("toggle_rule"))
async def cmd_toggle_rule(message: Message):
    if not is_allowed(message.from_user.id):
        return
    parts = (message.text or "").split()
    if len(parts) < 2 or not parts[1].isdigit():
        await message.answer("Использование: /toggle_rule <id>")
        return
    rule_id = int(parts[1])
    try:
        new_state = _shared_db().toggle_rule(rule_id)
        _shared_engine().invalidate_cache()
    except Exception as e:
        await message.answer(f"Ошибка: {e}")
        return
    if new_state is None:
        await message.answer(f"Правило #{rule_id} не найдено.")
        return
    emoji = "✅" if new_state else "❌"
    await message.answer(f"{emoji} Правило #{rule_id} {'активно' if new_state else 'выключено'}.")


@dp.message(Command("delete_rule"))
async def cmd_delete_rule(message: Message):
    if not is_allowed(message.from_user.id):
        return
    parts = (message.text or "").split()
    if len(parts) < 2 or not parts[1].isdigit():
        await message.answer("Использование: /delete_rule <id>")
        return
    rule_id = int(parts[1])
    kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="✅ Удалить", callback_data=f"del_rule:{rule_id}"),
        InlineKeyboardButton(text="❌ Отмена", callback_data="del_cancel"),
    ]])
    await message.answer(f"Удалить правило #{rule_id}?", reply_markup=kb)


# ── Главный обработчик текст/голос ───────────────────────────────
@dp.message(F.text | F.voice)
async def handle_message(message: Message):
    user_id = message.from_user.id
    if not is_allowed(user_id):
        return

    # Сброс контекста при паузе > 5 минут
    now_ts = datetime.now().timestamp()
    prev_ts = _last_message_time.get(user_id, 0)
    _last_message_time[user_id] = now_ts
    if prev_ts > 0 and (now_ts - prev_ts) > _CONTEXT_TTL:
        histories.pop(user_id, None)

    if message.voice:
        text = await transcribe_voice(message)
        if text.startswith("[Не удалось"):
            await message.answer(text)
            return
        text = _fix_asr(text)   # исправляем ошибки Whisper
        await message.answer(f"🎙 _{text}_", parse_mode="Markdown")
    else:
        text = message.text

    # Перехват «переделай» — reply на мой ответ
    if (message.reply_to_message
            and message.reply_to_message.from_user
            and message.reply_to_message.from_user.id == bot.id
            and text.lower().startswith("переделай")):
        await _do_reformat(message, text)
        return

    q = text.lower()

    # ── 0. Ожидающий коммерческий поиск (пользователь ответил на уточнение) ──
    if user_id in _pending_searches:
        prev_query, attempts = _pending_searches.pop(user_id)
        combined = prev_query + " " + text
        kind = detect_type(combined)
        clarification = get_clarification(combined, kind)
        if clarification and attempts < 2:
            _pending_searches[user_id] = (combined, attempts + 1)
            await message.answer(clarification)
            return
        elif clarification:
            # Исчерпали попытки — ищем с тем что есть
            pass
        stop = asyncio.Event()
        typing = asyncio.create_task(_keep_typing(message.chat.id, stop))
        try:
            result = await smart_search_and_answer(combined, ai_client, config.SEARCH_MODEL, kind)
        finally:
            stop.set(); typing.cancel()
            try: await typing
            except asyncio.CancelledError: pass
        await _send_smart_result(message, result)
        return

    # ── 1. Проектные заметки ──────────────────────────────────────
    proj_m = _PROJ_RE.match(text.strip())
    if proj_m:
        loop = asyncio.get_event_loop()
        result = await loop.run_in_executor(None, _save_proj_note_sync, proj_m.group(1), proj_m.group(2))
        await message.answer(result, parse_mode="Markdown")
        return

    # ── 2. Задача команде Губаха ──────────────────────────────────
    task_m = _TASK_RE.search(text)
    if task_m:
        result = await add_gubaha_task(task_m.group(1), task_m.group(2), task_m.group(3) or "")
        await message.answer(result, parse_mode="Markdown")
        return

    # ── 3. CRM запрос ─────────────────────────────────────────────
    crm_q = re.match(r"что по (.+?)\??$", q)
    if crm_q:
        await message.answer(crm_get(crm_q.group(1)), parse_mode="Markdown")
        return

    # ── 4. CRM авто-сохранение ────────────────────────────────────
    for trigger in _CRM_TRIGGERS:
        if trigger in q:
            # Простое авто-сохранение — имя после триггера
            idx = q.find(trigger)
            snippet = text[idx:idx+120]
            # Имя — первое слово с большой буквы после триггера
            name_m = re.search(r'\b([А-ЯЁ][а-яёА-ЯЁ]+)\b', snippet)
            if name_m:
                crm_save(name_m.group(1), text.strip())
            break

    # ── 5. Данные из таблиц ───────────────────────────────────────
    if any(k in q for k in _SHEETS_KEYWORDS):
        await message.answer("📊 Загружаю данные...")
        stop = asyncio.Event()
        typing = asyncio.create_task(_keep_typing(message.chat.id, stop))
        try:
            result = await query_sheets(text)
        except Exception as e:
            result = f"Ошибка загрузки данных: {e}"
        finally:
            stop.set(); typing.cancel()
            try: await typing
            except asyncio.CancelledError: pass
        await _safe_answer(message, result, parse_mode=None)
        return

    # ── 6. Почта — дайджест ───────────────────────────────────────
    if any(k in q for k in ["почта", "письм", "входящ", "пришло на почту"]):
        await message.answer("📬 Загружаю...")
        stop = asyncio.Event()
        typing = asyncio.create_task(_keep_typing(message.chat.id, stop))
        try:
            loop = asyncio.get_event_loop()
            result = await loop.run_in_executor(None, mail_digest_sync)
        finally:
            stop.set(); typing.cancel()
            try: await typing
            except asyncio.CancelledError: pass
        await _safe_answer(message, result)
        return

    # ── 7. Написать письмо ────────────────────────────────────────
    email_m = re.match(r"напиши(?:\s+письмо)?\s+(\w+)\s+(?:о|про|насчёт|по\s+поводу)?\s*(.+)", q)
    if email_m or ("отправь письмо" in q) or ("напиши письмо" in q):
        # Извлекаем получателя и тему
        if email_m:
            recipient = email_m.group(1)
            content = email_m.group(2)
        else:
            recipient = "себе"
            content = text
        try:
            result, kb = await compose_email(user_id, recipient, content[:50], content)
            await message.answer(result, reply_markup=kb, parse_mode="Markdown")
        except ValueError as e:
            await message.answer(f"❌ {e}")
        return

    # ── 8. Веб-поиск ─────────────────────────────────────────────
    # Продукты → прямые API магазинов; билеты/отели → Perplexity; общее → DuckDuckGo
    _is_commercial = COMMERCIAL_RE.search(text.strip())
    if _is_commercial or _SEARCH_RE.match(text.strip()) or any(k in q for k in ["расписание", "онлайн табло", "курс валют", "погода", "новости"]):
        if _is_commercial:
            kind = detect_type(text)
            if kind == "products":
                # Параллельный поиск через Perplexity site: по каждому магазину
                stop = asyncio.Event()
                typing = asyncio.create_task(_keep_typing(message.chat.id, stop))
                try:
                    shop_results = await search_all_stores(text, ai_client, config.SEARCH_MODEL)
                    response_text = format_shopping(text, shop_results)
                finally:
                    stop.set(); typing.cancel()
                    try: await typing
                    except asyncio.CancelledError: pass
                await _safe_answer(message, response_text, parse_mode=None)
                return
            clarification = get_clarification(text, kind)
            if clarification:
                # Не хватает данных — сохраняем запрос и задаём вопрос
                _pending_searches[user_id] = (text, 0)
                await message.answer(clarification)
                return
            stop = asyncio.Event()
            typing = asyncio.create_task(_keep_typing(message.chat.id, stop))
            try:
                result = await smart_search_and_answer(text, ai_client, config.SEARCH_MODEL, kind)
            finally:
                stop.set(); typing.cancel()
                try: await typing
                except asyncio.CancelledError: pass
            await _send_smart_result(message, result)
        else:
            stop = asyncio.Event()
            typing = asyncio.create_task(_keep_typing(message.chat.id, stop))
            try:
                result = await web_search_and_answer(text)
            finally:
                stop.set(); typing.cancel()
                try: await typing
                except asyncio.CancelledError: pass
            await _safe_answer(message, result)
        return

    # ── 9. AI-чат (fallback) ──────────────────────────────────────
    history = histories.setdefault(user_id, [])
    if not history:
        history.append({"role": "system", "content": _system_prompt()})
    history.append({"role": "user", "content": text})
    if len(history) > 22:
        histories[user_id] = history[:1] + history[-20:]
        history = histories[user_id]

    stop = asyncio.Event()
    typing = asyncio.create_task(_keep_typing(message.chat.id, stop))
    try:
        response = await run_llm(history)
    except Exception as e:
        response = f"Ошибка: {e}"
    finally:
        stop.set(); typing.cancel()
        try: await typing
        except asyncio.CancelledError: pass

    # LLM сигнализирует что нужен веб-поиск
    if response.strip() == "ПОИСК":
        if COMMERCIAL_RE.search(text):
            kind = detect_type(text)
            if kind == "products":
                stop2 = asyncio.Event()
                typing2 = asyncio.create_task(_keep_typing(message.chat.id, stop2))
                try:
                    shop_results = await search_all_stores(text, ai_client, config.SEARCH_MODEL)
                    response = format_shopping(text, shop_results)
                finally:
                    stop2.set(); typing2.cancel()
                    try: await typing2
                    except asyncio.CancelledError: pass
            else:
                clarification = get_clarification(text, kind)
                if clarification:
                    _pending_searches[user_id] = (text, 0)
                    response = clarification
                else:
                    stop2 = asyncio.Event()
                    typing2 = asyncio.create_task(_keep_typing(message.chat.id, stop2))
                    try:
                        smart_result = await smart_search_and_answer(text, ai_client, config.SEARCH_MODEL, kind)
                    finally:
                        stop2.set(); typing2.cancel()
                        try: await typing2
                        except asyncio.CancelledError: pass
                    await _send_smart_result(message, smart_result)
                    return
        else:
            stop2 = asyncio.Event()
            typing2 = asyncio.create_task(_keep_typing(message.chat.id, stop2))
            try:
                response = await web_search_and_answer(text)
            finally:
                stop2.set(); typing2.cancel()
                try: await typing2
                except asyncio.CancelledError: pass

    try:
        from rule_engine import apply_rules as _apply_rules
        response = await _apply_rules(response, text, "helper", ai_client)
    except Exception:
        pass

    history.append({"role": "assistant", "content": response})
    await _safe_answer(message, response)


# ══════════════════════════════════════════════════════════════════
# ДАЙДЖЕСТ HOTELIERРRO
# ══════════════════════════════════════════════════════════════════

def _hotelier_conn():
    _HOTELIER_DB.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(_HOTELIER_DB))
    conn.row_factory = sqlite3.Row
    return conn

def _init_hotelier_db():
    conn = _hotelier_conn()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS channel_posts (
            id INTEGER PRIMARY KEY,
            channel TEXT,
            message_id INTEGER,
            date TEXT,
            text TEXT,
            is_relevant INTEGER DEFAULT 0,
            topics TEXT,
            sent INTEGER DEFAULT 0,
            UNIQUE(channel, message_id)
        );
        CREATE TABLE IF NOT EXISTS bot_flags (
            key TEXT PRIMARY KEY,
            value TEXT
        );
    """)
    conn.commit()
    conn.close()

def _scrape_channel_page(before_id: int | None = None) -> tuple[list[dict], int | None]:
    """Парсит одну страницу t.me/s/HotelierPRO. Возвращает (посты, min_id на странице)."""
    url = "https://t.me/s/HotelierPRO"
    if before_id:
        url += f"?before={before_id}"
    headers = {"User-Agent": "Mozilla/5.0 (compatible; bot/1.0)"}
    try:
        r = requests.get(url, headers=headers, timeout=15)
        r.raise_for_status()
    except Exception as e:
        logging.warning(f"[hotelier] scrape error: {e}")
        return [], None
    soup = BeautifulSoup(r.text, "html.parser")
    posts = []
    min_id = None
    for wrap in soup.select(".tgme_widget_message_wrap"):
        msg_div = wrap.select_one(".tgme_widget_message")
        if not msg_div:
            continue
        data_post = msg_div.get("data-post", "")
        try:
            msg_id = int(data_post.split("/")[-1])
        except (ValueError, IndexError):
            continue
        time_tag = wrap.select_one("time[datetime]")
        if not time_tag:
            continue
        try:
            dt = datetime.fromisoformat(time_tag["datetime"]).astimezone(_MSK)
        except Exception:
            continue
        text_el = wrap.select_one(".tgme_widget_message_text")
        text = text_el.get_text("\n", strip=True) if text_el else ""
        if not text or len(text) < 30:
            continue
        posts.append({
            "message_id": msg_id,
            "date": dt.strftime("%Y-%m-%d %H:%M"),
            "date_dt": dt,
            "text": text[:2000],
        })
        if min_id is None or msg_id < min_id:
            min_id = msg_id
    return posts, min_id

async def _fetch_channel_posts(since_date: datetime, limit: int = 500) -> list[dict]:
    """Читает посты из @HotelierPRO через публичный веб-превью t.me/s/."""
    loop = asyncio.get_event_loop()
    all_posts: list[dict] = []
    before_id = None
    while len(all_posts) < limit:
        page_posts, min_id = await loop.run_in_executor(None, _scrape_channel_page, before_id)
        if not page_posts:
            break
        for p in page_posts:
            if p["date_dt"] >= since_date:
                all_posts.append({k: v for k, v in p.items() if k != "date_dt"})
        oldest_dt = min(p["date_dt"] for p in page_posts)
        if oldest_dt < since_date or min_id is None:
            break
        before_id = min_id
        await asyncio.sleep(1)  # вежливая пауза между запросами
    # убираем дубли, сортируем по дате
    seen = set()
    result = []
    for p in sorted(all_posts, key=lambda x: x["date"]):
        if p["message_id"] not in seen:
            seen.add(p["message_id"])
            result.append(p)
    return result

_KW_TEMA2 = re.compile(r"Пермск|Пермь|Губах|Прикамь|пермяк", re.IGNORECASE)

async def _filter_posts_by_topics(posts: list[dict]) -> dict[int, list[str]]:
    """Классификация через Claude — батчами по 15 постов."""
    result: dict[int, list[str]] = {}
    batch_size = 15
    for i in range(0, len(posts), batch_size):
        batch = posts[i:i + batch_size]
        numbered = "\n\n".join(
            f"ID={p['message_id']}\n{p['text'][:600]}"
            for p in batch
        )
        prompt = (
            "Ты классификатор постов из канала HotelierPRO.\n"
            "Для каждого поста выведи строго ОДНУ строку:\n"
            "ID=<число> T1  — пост содержит рыночную статистику: данные о загрузке/ценах/турпотоке/спросе/сезоне "
            "по отрасли или региону России (цифры, прогнозы, тренды рынка в целом).\n"
            "ID=<число> T2  — пост упоминает Пермский край (проекты, события, новости).\n"
            "ID=<число> T1 T2  — подходят обе.\n"
            "ID=<число> NO  — реклама продукта/сервиса, HR/вакансии, мотивация, новости конкретного отеля без рыночных цифр.\n\n"
            "Выводи ТОЛЬКО строки классификации, без пояснений.\n\n"
            + numbered
        )
        try:
            resp = await ai_client.chat.completions.create(
                model=getattr(config, "VISION_MODEL", config.MODEL),
                messages=[{"role": "user", "content": prompt}],
                max_tokens=400,
                temperature=0.0,
            )
            text = (resp.choices[0].message.content or "").strip()
            logging.warning(f"[hotelier] filter batch {i//batch_size}: {text}")
            for line in text.splitlines():
                m = re.match(r"ID=(\d+)\s+(.+)", line.strip())
                if not m:
                    continue
                msg_id = int(m.group(1))
                tags = m.group(2).upper()
                topics = []
                if "T1" in tags:
                    topics.append("статистика_отелей")
                if "T2" in tags or _KW_TEMA2.search(
                    next((p["text"] for p in batch if p["message_id"] == msg_id), "")
                ):
                    topics.append("пермский_край")
                if topics:
                    result[msg_id] = topics
        except Exception as e:
            logging.warning(f"[hotelier] filter batch error: {e}")
    return result

async def hotelier_fetch_and_store():
    """Сбор новых постов каждые 4 часа."""
    try:
        conn = _hotelier_conn()
        row = conn.execute("SELECT MAX(date) as last FROM channel_posts").fetchone()
        conn.close()
        if row and row["last"]:
            since = datetime.strptime(row["last"], "%Y-%m-%d %H:%M").replace(tzinfo=_MSK)
        else:
            since = datetime.now(_MSK) - timedelta(hours=4)
        posts = await _fetch_channel_posts(since_date=since)
        if not posts:
            logging.info("[hotelier] no new posts")
            return
        relevance = await _filter_posts_by_topics(posts)
        if not relevance:
            logging.info(f"[hotelier] fetched {len(posts)} posts, 0 relevant")
            return
        conn = _hotelier_conn()
        saved = 0
        for p in posts:
            if p["message_id"] in relevance:
                topics = json.dumps(relevance[p["message_id"]], ensure_ascii=False)
                try:
                    conn.execute(
                        "INSERT OR IGNORE INTO channel_posts (channel, message_id, date, text, is_relevant, topics) "
                        "VALUES (?, ?, ?, ?, 1, ?)",
                        ("HotelierPRO", p["message_id"], p["date"], p["text"], topics),
                    )
                    saved += 1
                except Exception:
                    pass
        conn.commit()
        conn.close()
        logging.info(f"[hotelier] fetched {len(posts)} posts, {saved} relevant saved")
    except Exception as e:
        logging.error(f"[hotelier] fetch_and_store error: {e}")

async def _hotelier_send(uid: int, text: str):
    """Отправляет сообщение с разбивкой на части и fallback на plain text."""
    chunk_size = 3800
    parts = [text[i:i + chunk_size] for i in range(0, len(text), chunk_size)]
    for part in parts:
        try:
            await bot.send_message(uid, part, parse_mode="Markdown")
        except Exception:
            try:
                await bot.send_message(uid, part, parse_mode=None)
            except Exception as e:
                logging.warning(f"[hotelier] send failed: {e}")

def _hotelier_format_post(row) -> str:
    """Форматирует пост: оригинальный текст + ссылка на источник."""
    topics = json.loads(row["topics"] or "[]")
    tag = ""
    if "статистика_отелей" in topics and "пермский_край" in topics:
        tag = "📊🏔 "
    elif "статистика_отелей" in topics:
        tag = "📊 "
    elif "пермский_край" in topics:
        tag = "🏔 "
    msg_id = row["message_id"]
    date_str = row["date"]
    text = row["text"]
    link = f"https://t.me/HotelierPRO/{msg_id}"
    return f"{tag}{date_str}\n\n{text}\n\n{link}"

async def _hotelier_send_posts(uid: int, rows: list, header: str = ""):
    """Отправляет каждый пост отдельным сообщением."""
    if header:
        try:
            await bot.send_message(uid, header, parse_mode=None)
        except Exception as e:
            logging.warning(f"[hotelier] header send failed: {e}")
        await asyncio.sleep(0.3)
    for row in rows:
        text = _hotelier_format_post(row)
        # Telegram лимит 4096, режем если надо
        if len(text) > 4000:
            text = text[:3950] + f"...\n\nhttps://t.me/HotelierPRO/{row['message_id']}"
        try:
            await bot.send_message(uid, text, parse_mode=None, disable_web_page_preview=True)
        except Exception as e:
            logging.warning(f"[hotelier] post send failed (id={row['message_id']}): {e}")
        await asyncio.sleep(0.3)

async def hotelier_digest():
    """Ежедневный дайджест в 10:00 МСК."""
    try:
        since = datetime.now(_MSK) - timedelta(hours=24)
        conn = _hotelier_conn()
        rows = conn.execute(
            "SELECT id, message_id, date, text, topics FROM channel_posts "
            "WHERE is_relevant=1 AND sent=0 AND date >= ? "
            "ORDER BY date",
            (since.strftime("%Y-%m-%d %H:%M"),),
        ).fetchall()
        if not rows:
            conn.close()
            logging.info("[hotelier] digest: no new relevant posts")
            return
        header = f"HotelierPRO — {datetime.now(_MSK).strftime('%d.%m.%Y')} ({len(rows)} постов)"
        ids = [r["id"] for r in rows]
        for uid in config.ALLOWED_USER_IDS:
            await _hotelier_send_posts(uid, rows, header)
        conn.execute(f"UPDATE channel_posts SET sent=1 WHERE id IN ({','.join('?' * len(ids))})", ids)
        conn.commit()
        conn.close()
        logging.info(f"[hotelier] digest sent, {len(rows)} posts")
    except Exception as e:
        logging.error(f"[hotelier] digest error: {e}")

async def hotelier_initial_load():
    """Однократная загрузка истории за 2 месяца при первом старте."""
    await asyncio.sleep(5)
    try:
        conn = _hotelier_conn()
        flag = conn.execute("SELECT value FROM bot_flags WHERE key='hotelier_initial_done'").fetchone()
        unsent = conn.execute(
            "SELECT id, message_id, date, text, topics FROM channel_posts "
            "WHERE is_relevant=1 AND sent=0 ORDER BY date"
        ).fetchall()
        conn.close()

        if flag and not unsent:
            return

        if not flag:
            logging.info("[hotelier] initial load started, fetching 2 months...")
            since = datetime.now(_MSK) - timedelta(days=62)
            posts = await _fetch_channel_posts(since_date=since, limit=2000)
            logging.info(f"[hotelier] initial: got {len(posts)} total posts")
            if not posts:
                conn = _hotelier_conn()
                conn.execute("INSERT OR REPLACE INTO bot_flags VALUES ('hotelier_initial_done', '1')")
                conn.commit()
                conn.close()
                return
            relevance = await _filter_posts_by_topics(posts)
            logging.info(f"[hotelier] initial: {len(relevance)} relevant")
            conn = _hotelier_conn()
            for p in posts:
                if p["message_id"] in relevance:
                    topics = json.dumps(relevance[p["message_id"]], ensure_ascii=False)
                    conn.execute(
                        "INSERT OR IGNORE INTO channel_posts (channel, message_id, date, text, is_relevant, topics) "
                        "VALUES (?, ?, ?, ?, 1, ?)",
                        ("HotelierPRO", p["message_id"], p["date"], p["text"], topics),
                    )
            conn.execute("INSERT OR REPLACE INTO bot_flags VALUES ('hotelier_initial_done', '1')")
            conn.commit()
            conn.close()
            conn = _hotelier_conn()
            unsent = conn.execute(
                "SELECT id, message_id, date, text, topics FROM channel_posts "
                "WHERE is_relevant=1 AND sent=0 ORDER BY date"
            ).fetchall()
            conn.close()

        if not unsent:
            logging.info("[hotelier] initial: no relevant posts found")
            return

        since_str = (datetime.now(_MSK) - timedelta(days=62)).strftime("%d.%m")
        now_str = datetime.now(_MSK).strftime("%d.%m.%Y")
        header = f"HotelierPRO — сводка за {since_str}–{now_str} ({len(unsent)} постов)"
        ids = [r["id"] for r in unsent]
        for uid in config.ALLOWED_USER_IDS:
            await _hotelier_send_posts(uid, unsent, header)
        conn = _hotelier_conn()
        conn.execute(f"UPDATE channel_posts SET sent=1 WHERE id IN ({','.join('?' * len(ids))})", ids)
        conn.commit()
        conn.close()
        logging.info(f"[hotelier] initial load complete, sent {len(unsent)} posts")
    except Exception as e:
        logging.error(f"[hotelier] initial_load error: {e}")


# ══════════════════════════════════════════════════════════════════
# ЗАПУСК
# ══════════════════════════════════════════════════════════════════
async def main():
    _init_hotelier_db()
    scheduler.add_job(hotelier_fetch_and_store, "interval", hours=4, id="hotelier_fetch", replace_existing=True)
    scheduler.add_job(hotelier_digest, "cron", hour=10, minute=0, id="hotelier_digest", replace_existing=True)
    scheduler.start()
    asyncio.create_task(hotelier_initial_load())
    print(f"[{datetime.now(_MSK).strftime('%H:%M:%S')} МСК] Помощник запущен.")
    await dp.start_polling(bot, allowed_updates=["message", "callback_query"])

if __name__ == "__main__":
    asyncio.run(main())
